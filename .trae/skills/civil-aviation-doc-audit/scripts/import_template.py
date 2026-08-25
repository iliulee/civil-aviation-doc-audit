"""
import_template.py — 从电子版表格（.xlsx / .docx）导入表格式样

把电子版表格解析成「表格式样」JSON（前缀区固定行 + 数字区每列定义），
供 data-editor.html 的「表格式样 → 导入 JSON」按钮加载。

用法：
    python scripts/import_template.py <电子表格文件> [--out <输出.json>] [--sheet <工作表>] [--header-row <行号(1起)>]

说明：
    - 自动检测表头行（含序号/桩号/日期/沉管/拔管/桩长等关键词的行）。
    - 表头行及以上非空行 → 前缀区（prefix_rows）。
    - 表头行列名 → 数字区每列 name；从数据区取首个非空值作为样例填进 styles（供人工补成格式模式）。
    - --header-row 可手动指定表头行，跳过自动检测。
"""
import argparse
import json
import os
import sys
from datetime import datetime

# 表头关键词：命中即视为表头候选
HEADER_KEYWORDS = [
    '序号', '桩号', '施工日期', '日期', '时间', '沉管', '拔管', '桩长', '桩径',
    '高程', '桩顶', '桩底', '直径', '电流', '充盈', '竖直', '垂直', '灌入',
    '反插', '复打', '设计', '实际', '密实', '深度', '料', '备注', '桩体',
]

SAMPLE_COL_LIMIT = 80  # styles 样例值最大长度


def _cell_text(v):
    """单元格值 → 字符串；None/空返回 ''。"""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _header_score(row):
    """估算一行作为表头的置信度：命中关键词数 + 短文本填充度。"""
    kw = 0
    filled = 0
    for v in row:
        t = _cell_text(v)
        if not t:
            continue
        filled += 1
        if len(t) <= 12 and any(k in t for k in HEADER_KEYWORDS):
            kw += 1
    return kw + filled * 0.1


def _detect_header_row(rows, force_row=None):
    """返回表头行索引；force_row 为 1 起行号。找不到返回 None。"""
    if force_row:
        idx = force_row - 1
        if 0 <= idx < len(rows):
            return idx
    best_i, best_s = -1, -1
    for i, row in enumerate(rows):
        s = _header_score(row)
        if s > best_s:
            best_s, best_i = s, i
    return best_i if best_s >= 1.0 else None


def _read_xlsx(path, sheet=None):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_text(v) for v in row])
    return rows


def _read_docx(path):
    import docx
    d = docx.Document(path)
    rows = []
    for table in d.tables:
        for row in table.rows:
            rows.append([_cell_text(c.text) for c in row.cells])
        if rows:
            break  # 只取第一个表格
    return rows


def build_style(rows, header_idx):
    """由二维行列表 + 表头行索引，生成 table_style 结构。"""
    prefix_rows = []
    for i in range(header_idx):
        text = ' / '.join(t for t in rows[i] if t)
        if text:
            prefix_rows.append({'text': text})

    header = rows[header_idx]
    # 列边界：非空表头单元格占位起始列
    col_starts = []
    for c, t in enumerate(header):
        if t:
            col_starts.append(c)
    # 数字列：按表头单元格生成
    number_cols = []
    for c in col_starts:
        name = header[c]
        # 数据区取该列首个非空值作为样例
        sample = ''
        for r in range(header_idx + 1, len(rows)):
            v = rows[r][c] if c < len(rows[r]) else ''
            if v:
                sample = v[:SAMPLE_COL_LIMIT]
                break
        number_cols.append({'name': name, 'styles': sample, 'remark': ''})

    return {'prefix_rows': prefix_rows, 'number_cols': number_cols}


def main():
    ap = argparse.ArgumentParser(description='从电子版表格导入表格式样')
    ap.add_argument('file', help='电子表格路径（.xlsx / .docx）')
    ap.add_argument('--out', default=None, help='输出 JSON 路径；缺省写同目录 <原名>.table_style.json')
    ap.add_argument('--sheet', default=None, help='Excel 工作表名（缺省用活动表）')
    ap.add_argument('--header-row', type=int, default=0, help='表头行号(1起)，跳过自动检测')
    args = ap.parse_args()

    if not os.path.isfile(args.file):
        sys.exit(f'文件不存在: {args.file}')

    ext = os.path.splitext(args.file)[1].lower()
    if ext == '.xlsx':
        rows = _read_xlsx(args.file, args.sheet)
    elif ext == '.docx':
        rows = _read_docx(args.file)
    else:
        sys.exit(f'不支持的文件类型: {ext}（仅支持 .xlsx / .docx）')

    header_idx = _detect_header_row(rows, args.header_row)
    if header_idx is None:
        sys.exit('未能自动识别表头行，请用 --header-row <行号> 手动指定')

    style = build_style(rows, header_idx)

    out = {
        'schema_version': '1.0',
        'source_file': os.path.basename(args.file),
        'derived_at': datetime.now().replace(microsecond=0).isoformat(),
        'header_row': header_idx + 1,
        'table_style': style,
    }

    out_path = args.out or (
        os.path.splitext(args.file)[0] + '.table_style.json'
    )
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f'表头行: {header_idx + 1}')
    print(f'前缀区: {len(style["prefix_rows"])} 行')
    print(f'数字区: {len(style["number_cols"])} 列')
    print(f'已写出: {out_path}')


if __name__ == '__main__':
    main()