# -*- coding: utf-8 -*-
"""
数据底座建立脚本（build_foundation.py）
======================================

Phase 1 核心脚本：扫描项目文件夹 → 文件分类 → OCR/PDF 提取 → 结构化 JSON →
数据质量检测 → 混淆检测 → index.json 总索引 → 复制 Web 模板。

用法：
    python scripts/build_foundation.py <项目文件夹路径> \
        --engine <auto|vision|paddle> \
        --incremental \
        --out <数据底座目录名，默认"数据底座"> \
        --preconditions <前置信息JSON文件路径> \
        --expected-rows <预期行数JSON文件路径，可选>

约束：
    - 不进入正式审核（Phase 1 结束后停止）。
    - 不引入数据库、后端服务、云端依赖（OCR 引擎选择 vision 时依赖用户本地 API Key）。
    - index.json 是唯一真相源，任何数据变更都会更新它。
"""

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# 允许 import 同目录下的 run_audit.py
SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))

from run_audit import sniff_document  # noqa: E402
from audit_config import assign_subdivision_to_document, get_subdivision_info  # noqa: E402

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False


# ========== 路径常量 ==========
OCR_IMAGE_PY = SCRIPT_DIR / "ocr_image.py"
EXTRACT_PDF_PY = SCRIPT_DIR / "extract_pdf.py"
DQ_CHECK_PY = SCRIPT_DIR / "data_quality_check.py"
CONFUSION_CHECK_PY = SCRIPT_DIR / "ocr_confusion_check.py"
TEMPLATES_DIR = SKILL_DIR / "templates"

# ========== 直接导入 OCR 函数（避免子进程崩溃） ==========
try:
    from ocr_image import ocr_pdf as _ocr_pdf_direct
    _OCR_DIRECT_AVAILABLE = True
except ImportError:
    _OCR_DIRECT_AVAILABLE = False

# 页面平均行数（用于未提供 expected-rows 时的推断）
DEFAULT_ROWS_PER_PDF_PAGE = 20
DEFAULT_ROWS_PER_IMAGE_PAGE = 15

# 行数校验阈值：实际行数低于预期的 80% 时触发重试
ROW_COUNT_THRESHOLD = 0.8

# 默认纳入扫描的被审核资料扩展名（--include-all 可关闭此过滤）
DEFAULT_ALLOWED_EXTENSIONS = {
    ".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp",
    ".txt", ".md", ".doc", ".docx", ".xls", ".xlsx",
}


# ========== 专业/资料分类规则 ==========
# v7.2 C1: PROFESSIONAL_RULES 硬编码表已删除，改为运行时聚合三真相源
# 真相源：references/classification-terms.json + rules trigger_when.doc_type + FIELD_ALIAS_MAP
# 聚合函数见 aggregate_classification_terms()，在 classify_file 中调用
# v7.2 P1-02: REFERENCE_KEYWORDS/GENERIC_KEYWORDS 迁移到 classification-terms.json

def _load_reference_keywords() -> list:
    """从 classification-terms.json 加载依据文件关键词"""
    terms_file = SKILL_DIR / "references" / "classification-terms.json"
    terms_data = load_json(terms_file) or {}
    return terms_data.get("reference_keywords", [])

def _load_generic_keywords() -> dict:
    """从 classification-terms.json 加载通用资料关键词"""
    terms_file = SKILL_DIR / "references" / "classification-terms.json"
    terms_data = load_json(terms_file) or {}
    raw = terms_data.get("generic_keywords", {})
    # 转换为 (prof, sub, doc) 元组格式，保持向后兼容
    result = {}
    for kw, info in raw.items():
        result[kw] = (info.get("professional", ""), info.get("subdivision", ""), info.get("doc_type", ""))
    return result

# v7.2 C5: 桩基表头关键词——从 FIELD_ALIAS_MAP 反向聚合（单一真相源）
# 额外的 OCR 专用别名（FIELD_ALIAS_MAP 中没有的）在此补充
_OCR_EXTRA_ALIASES: Dict[str, List[str]] = {
    "pile_no": ["序号/桩", "序号"],
    "start_time": ["沉管开始"],
    "end_time": ["拔管结束"],
    "sink_time": ["沉管时长", "沉管开始时间"],
    "pull_time": ["拔管时长", "拔管结束时间"],
}

_PILE_SLOT_SCHEMA_CACHE: Optional[Dict[str, Any]] = None


def get_pile_slot_schema() -> Dict[str, Any]:
    """v7.2 C5: 从 FIELD_ALIAS_MAP 反向聚合桩基槽位 schema。
    v7.2 V-70: 同时读取 header-aliases.json 中 status=active 的候选别名（表头自成长）。
    返回: {slot: {aliases: [...], data_type: str, num_range_hint: str}}
    """
    global _PILE_SLOT_SCHEMA_CACHE
    if _PILE_SLOT_SCHEMA_CACHE is not None:
        return _PILE_SLOT_SCHEMA_CACHE

    schema: Dict[str, Any] = {}

    # 从 FIELD_ALIAS_MAP 反向聚合
    try:
        from rule_engine import FIELD_ALIAS_MAP  # noqa: E402
        for cn_name, en_slot in FIELD_ALIAS_MAP.items():
            schema.setdefault(en_slot, {"aliases": [], "data_type": "str", "num_range_hint": ""})
            schema[en_slot]["aliases"].append(cn_name)
    except ImportError:
        pass

    # 补充 OCR 专用别名
    for slot, extra_aliases in _OCR_EXTRA_ALIASES.items():
        schema.setdefault(slot, {"aliases": [], "data_type": "str", "num_range_hint": ""})
        for alias in extra_aliases:
            if alias not in schema[slot]["aliases"]:
                schema[slot]["aliases"].append(alias)

    # v7.2 V-70: 读取 header-aliases.json 中 status=active 的候选别名（表头自成长回流）
    header_aliases_file = SKILL_DIR / "references" / "header-aliases.json"
    header_data = load_json(header_aliases_file) or {}
    for slot, slot_data in header_data.get("slots", {}).items():
        schema.setdefault(slot, {"aliases": [], "data_type": "str", "num_range_hint": ""})
        for alias in slot_data.get("aliases", []):
            if alias not in schema[slot]["aliases"]:
                schema[slot]["aliases"].append(alias)
        for cand in slot_data.get("candidates", []):
            if cand.get("status") == "active" and cand.get("alias") not in schema[slot]["aliases"]:
                schema[slot]["aliases"].append(cand["alias"])

    # 数据类型提示
    numeric_slots = {"design_length", "diameter", "bottom_elev", "top_elev", "actual_length",
                     "current", "re_penetration", "volume", "filling_coeff", "verticality"}
    time_slots = {"start_time", "end_time", "sink_time", "pull_time"}
    for slot in schema:
        if slot in numeric_slots:
            schema[slot]["data_type"] = "number"
        elif slot in time_slots:
            schema[slot]["data_type"] = "time"

    _PILE_SLOT_SCHEMA_CACHE = schema
    return schema


def _get_pile_header_keywords() -> Dict[str, List[str]]:
    """v7.2 C5: 从槽位 schema 获取表头关键词映射（替代硬编码 PILE_HEADER_KEYWORDS）。"""
    schema = get_pile_slot_schema()
    return {slot: data["aliases"] for slot, data in schema.items()}


# 碎石桩表头关键词 → 字段名（v7.2 C5: 从 FIELD_ALIAS_MAP 运行时聚合，不再硬编码）
# 保留变量名兼容已有代码，但值在首次使用时从 get_pile_slot_schema() 聚合
PILE_HEADER_KEYWORDS: Dict[str, List[str]] = {}  # 延迟初始化，使用前调用 _get_pile_header_keywords()

PILE_FIELDS = [
    "pile_no", "design_length", "diameter", "bottom_elev", "top_elev",
    "actual_length", "current", "re_penetration", "volume", "filling_coeff",
    "verticality", "start_time", "end_time", "sink_time", "pull_time", "remark",
]


# ========== 工具函数 ==========
def now_iso() -> str:
    """返回当前时间的 ISO 8601 字符串。"""
    return datetime.now().isoformat(timespec="seconds")


def generate_default_preconditions(project_path: Path) -> Dict[str, Any]:
    """未提供前置信息文件时，生成显式默认值并保存到输出目录供追溯。"""
    return {
        "stage": "分部分项验收",
        "nature": "扫描件",
        "scope": "全量审核",
        "ocr_engine": "auto",
        "check_signatures": False,
        "special_notes": "由 build_foundation.py 自动生成默认前置信息，未经过人工确认",
        "excluded_files": [],
        "expected_rows": {},
    }


def load_json(path: Path) -> Any:
    """读取 JSON 文件，不存在时返回 None。"""
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: Any) -> None:
    """以 UTF-8 写入 JSON，保持中文可读。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def safe_filename(name: str) -> str:
    """去除 Windows 非法字符，用于生成文件/目录名。"""
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


def run_subprocess(cmd: List[str]) -> Tuple[int, str, str]:
    """运行子进程命令，返回 (returncode, stdout, stderr)。"""
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        return result.returncode, result.stdout, result.stderr
    except Exception as e:
        return -1, "", str(e)


def _compute_file_hash(file_path: Path) -> str:
    """计算文件的 SHA256 哈希值（用于增量更新检测）。

    对大文件采用分块读取，避免内存溢出。
    """
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


# ========== 文件扫描与分类 ==========
def is_temp_or_hidden(rel_path: Path) -> bool:
    """判断相对路径是否包含隐藏文件/目录或临时文件。"""
    parts = rel_path.parts
    for part in parts:
        if part.startswith("."):
            return True
    name = rel_path.name
    if name.startswith("~") or name.startswith("."):
        return True
    if name.lower().endswith((".tmp", ".bak", ".log", ".db", ".cache")):
        return True
    return False


def scan_files(
    project_path: Path,
    out_name: str,
    include_all: bool = False,
) -> List[Path]:
    """
    扫描项目文件夹下所有文件，排除数据底座目录自身、输出文件、隐藏/临时文件。
    默认仅保留资料类扩展名，--include-all 时扫描所有扩展名。
    返回文件相对 Path 列表。
    """
    files: List[Path] = []
    output_root_files = {"data-editor.html", "项目总览.html", "审核报告.html", "tokens.css"}
    # 排除任意层级的输出目录：用户指定的 out_name 以及默认的"数据底座"
    output_dir_names = {out_name, "数据底座"}
    for p in project_path.rglob("*"):
        if not p.is_file():
            continue
        rel = p.relative_to(project_path)
        # 排除输出目录（任意层级）
        if any(part in output_dir_names for part in rel.parts[:-1]):
            continue
        # 排除 Web/报告产物（任意层级）
        if rel.name in output_root_files:
            continue
        if is_temp_or_hidden(rel):
            continue
        # 默认扩展名过滤
        if not include_all and p.suffix.lower() not in DEFAULT_ALLOWED_EXTENSIONS:
            continue
        files.append(rel)
    return sorted(files)


def _read_text_preview(abs_path: Path) -> str:
    """v7.2 C1: 轻量读取文件前几行文本摘要（供 LLM 分类语义判定）。
    失败安全：任何读取异常都返回空字符串，绝不影响 build 主流程。
    """
    try:
        suffix = abs_path.suffix.lower()
        lines: List[str] = []
        if suffix in (".txt", ".md", ".log"):
            with open(abs_path, "r", encoding="utf-8", errors="replace") as f:
                for _ in range(5):
                    line = f.readline()
                    if not line:
                        break
                    lines.append(line.strip())
        elif suffix in (".pdf",) and HAS_PYMUPDF:
            try:
                import fitz  # noqa: E402
                with fitz.open(str(abs_path)) as doc:
                    if doc.page_count > 0:
                        text = doc[0].get_text()
                        lines = [ln.strip() for ln in text.splitlines() if ln.strip()][:5]
            except Exception:
                return ""
        elif suffix in (".xlsx", ".xls"):
            try:
                import openpyxl  # noqa: E402
                wb = openpyxl.load_workbook(str(abs_path), read_only=True, data_only=True)
                ws = wb.worksheets[0]
                for row in ws.iter_rows(max_row=4, values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        lines.append(" ".join(cells))
                wb.close()
            except Exception:
                return ""
        return "\n".join(lines)[:300]
    except Exception:
        return ""


def _has_formal_records(rel_files: List[Path]) -> bool:
    """
    判断文件列表中是否包含正式施工记录/检验批等应逐行审核的资料。
    用于施工日志自动降级为依据文件。
    """
    formal_keywords = ["施工记录", "检验批", "隐蔽工程", "碎石桩", "CFG桩", "桩基"]
    for rel in rel_files:
        name = rel.name
        # 施工日志自身不算
        if "施工日志" in name:
            continue
        if any(kw in name for kw in formal_keywords):
            return True
    return False


# ========== v7.2 C1: 分类词表运行时聚合 ==========
_AGGREGATED_TERMS: Optional[Dict[str, List[Dict[str, str]]]] = None


def aggregate_classification_terms() -> Dict[str, List[Dict[str, str]]]:
    """v7.2 C1: 从三真相源运行时聚合分类词表。

    真相源：
    1. references/classification-terms.json（结构化术语表）
    2. rules/**/*.json 的 trigger_when.doc_type + applies_to.professional
    3. rule_engine.py 的 FIELD_ALIAS_MAP（桩基字段名→场道工程弱信号）

    返回: {keyword: [{prof, weight, source}, ...]}
    """
    terms: Dict[str, List[Dict[str, str]]] = {}

    # Source 1: classification-terms.json
    terms_file = SKILL_DIR / "references" / "classification-terms.json"
    terms_data = load_json(terms_file) or {}
    for prof, data in terms_data.get("terms", {}).items():
        for kw in data.get("core", []):
            terms.setdefault(kw, []).append({"prof": prof, "weight": "core", "source": "terms"})
        for kw in data.get("weak", []):
            terms.setdefault(kw, []).append({"prof": prof, "weight": "weak", "source": "terms"})
    # candidates（自成长候选词条）
    for cand in terms_data.get("candidates", []):
        if cand.get("status") == "active":
            prof = cand.get("prof", "")
            kw = cand.get("keyword", "")
            if prof and kw:
                terms.setdefault(kw, []).append({"prof": prof, "weight": "core", "source": "candidate"})

    # Source 2: rules trigger_when.doc_type + applies_to.professional
    rules_dir = SKILL_DIR / "rules"
    if rules_dir.exists():
        for rule_file in rules_dir.rglob("*.json"):
            if rule_file.name in ("registry.json", "rule-schema.json", "registry-schema.json"):
                continue
            try:
                rule = json.loads(rule_file.read_text(encoding="utf-8"))
                doc_types = rule.get("trigger_when", {}).get("doc_type", [])
                profs = rule.get("applies_to", {}).get("professional", [])
                for dt in doc_types:
                    if dt == "*":
                        continue
                    for prof in profs:
                        if prof == "*":
                            continue
                        terms.setdefault(dt, []).append({"prof": prof, "weight": "core", "source": "rules"})
            except (json.JSONDecodeError, KeyError):
                continue

    # Source 3: FIELD_ALIAS_MAP（桩基字段名 → 场道工程弱信号）
    try:
        from rule_engine import FIELD_ALIAS_MAP  # noqa: E402
        pile_keywords = ["桩", "高程", "灌入", "电流", "沉管", "拔管", "充盈", "反插"]
        for cn_name in FIELD_ALIAS_MAP:
            if any(kw in cn_name for kw in pile_keywords):
                terms.setdefault(cn_name, []).append(
                    {"prof": "01_场道工程", "weight": "weak", "source": "field_alias"}
                )
    except ImportError:
        pass

    return terms


def get_aggregated_terms() -> Dict[str, List[Dict[str, str]]]:
    """获取聚合分类词表（带缓存）。"""
    global _AGGREGATED_TERMS
    if _AGGREGATED_TERMS is None:
        _AGGREGATED_TERMS = aggregate_classification_terms()
    return _AGGREGATED_TERMS


def classify_file(
    rel_path: Path,
    excluded_set: set,
    has_formal_records: bool = False,
    text_preview: str = "",
) -> Tuple[str, Optional[str], Optional[str], Optional[str], str, float]:
    """
    v7.2 C1: 对单个文件进行分类（三级判定：关键词聚合 → LLM语义 → 人工确认）。

    返回：
        (classification, professional, subcategory, doc_type, classification_source, classification_confidence)
        classification ∈ {"audited_files", "reference_files", "excluded_files"}
        classification_source ∈ {"terms", "rules", "field_alias", "generic_keywords", "reference_keywords", "weak", "weak_multi", "multi_hit", "llm", "default", "excluded"}
        classification_confidence ∈ [0.0, 1.0]
    """
    name = rel_path.name

    # 1. 用户明确排除
    if name in excluded_set or str(rel_path) in excluded_set:
        return "excluded_files", None, None, None, "excluded", 1.0

    # 2. 依据文件（设计变更类）
    for kw in _load_reference_keywords():
        if kw in name:
            return "reference_files", "依据文件", "设计依据", "设计变更文件", "reference_keywords", 1.0

    # 3. 通用资料
    for kw, (prof, sub, doc) in _load_generic_keywords().items():
        if kw in name:
            if kw == "施工日志" and has_formal_records:
                return "reference_files", "依据文件", "施工日志", "施工日志", "generic_keywords", 1.0
            return "audited_files", prof, sub, doc, "generic_keywords", 1.0

    # 4. v7.2 C1: 五大专业匹配（运行时聚合词表，替代硬编码 PROFESSIONAL_RULES）
    aggregated = get_aggregated_terms()
    hits: Dict[str, List[Dict[str, str]]] = {}  # prof → list of hit entries
    for kw, entries in aggregated.items():
        if kw in name:
            for entry in entries:
                prof = entry["prof"]
                hits.setdefault(prof, []).append(entry)

    if hits:
        strong_profs = [(p, e) for p, e in hits.items() if any(x["weight"] == "core" for x in e)]
        weak_profs = [(p, e) for p, e in hits.items() if not any(x["weight"] == "core" for x in e)]

        if len(strong_profs) == 1:
            # 单专业 core 强命中 → 直接分类，不上 LLM（省成本）
            prof, entries = strong_profs[0]
            sub, doc = detect_subcategory_and_doctype(name, prof)
            source = next((e["source"] for e in entries if e["weight"] == "core"), "aggregated")
            return "audited_files", prof, sub, doc, source, 0.9
        elif len(strong_profs) > 1:
            # 多专业 core 命中 → LLM 语义判定，失败则取首个标待确认
            llm_result = _try_llm_classify(name, text_preview)
            if llm_result:
                prof, sub, doc = _apply_llm_result(name, llm_result)
                return "audited_files", prof, sub, doc, "llm", llm_result["confidence"]
            prof, entries = strong_profs[0]
            sub, doc = detect_subcategory_and_doctype(name, prof)
            return "audited_files", prof, sub, doc, "multi_hit", 0.6
        elif len(weak_profs) == 1:
            # 单专业 weak 命中 → LLM 语义判定，失败则取该专业标待确认
            llm_result = _try_llm_classify(name, text_preview)
            if llm_result:
                prof, sub, doc = _apply_llm_result(name, llm_result)
                return "audited_files", prof, sub, doc, "llm", llm_result["confidence"]
            prof, entries = weak_profs[0]
            sub, doc = detect_subcategory_and_doctype(name, prof)
            return "audited_files", prof, sub, doc, "weak", 0.6
        elif len(weak_profs) > 1:
            # 多专业 weak 命中 → LLM 语义判定，失败则取首个标待确认
            llm_result = _try_llm_classify(name, text_preview)
            if llm_result:
                prof, sub, doc = _apply_llm_result(name, llm_result)
                return "audited_files", prof, sub, doc, "llm", llm_result["confidence"]
            prof, entries = weak_profs[0]
            sub, doc = detect_subcategory_and_doctype(name, prof)
            return "audited_files", prof, sub, doc, "weak_multi", 0.4

    # 5. 无任何关键词命中 → LLM 语义判定，失败则默认通用资料（标待确认）
    llm_result = _try_llm_classify(name, text_preview)
    if llm_result:
        prof, sub, doc = _apply_llm_result(name, llm_result)
        return "audited_files", prof, sub, doc, "llm", llm_result["confidence"]

    return "audited_files", "通用资料", "其他", "其他资料", "default", 0.3


def _try_llm_classify(name: str, text_preview: str) -> Optional[Dict[str, Any]]:
    """v7.2 C1: 调用 LLM 语义判定，失败/不可用返回 None（调用方降级）。"""
    try:
        from llm_client import classify_document  # noqa: E402
    except ImportError:
        return None
    try:
        return classify_document(name, text_preview)
    except Exception:
        # 失败安全：LLM 异常绝不影响 build 主流程
        return None


def _apply_llm_result(name: str, llm_result: Dict[str, Any]) -> Tuple[str, str, str]:
    """应用 LLM 分类结果，返回 (professional, subcategory, doc_type)。"""
    prof = llm_result.get("professional", "通用资料")
    if prof == "通用资料":
        return prof, "其他", "其他资料"
    sub, doc = detect_subcategory_and_doctype(name, prof)
    return prof, sub, doc


def detect_subcategory_and_doctype(name: str, professional: str) -> Tuple[str, str]:
    """根据文件名关键词推测 subcategory 与 doc_type。"""
    lower = name.lower()

    if professional == "01_场道工程":
        if "碎石桩" in name:
            return "施工记录", "碎石桩施工记录"
        if "cfg" in lower:
            return "施工记录", "CFG桩施工记录"
        if "桩" in name:
            return "施工记录", "桩基施工记录"
        if "土方" in name:
            return "施工记录", "土方工程"
        if "换填" in name:
            return "施工记录", "换填施工记录"
        if "混凝土" in name:
            return "施工记录", "混凝土工程"
        if any(kw in name for kw in ["跑道", "滑行道"]):
            return "施工记录", "场道施工记录"
        return "施工记录", "场道工程"

    if professional == "02_空管工程":
        if "vor" in lower:
            return "施工记录", "VOR台施工记录"
        if "dme" in lower:
            return "施工记录", "DME台施工记录"
        if "ils" in lower:
            return "施工记录", "ILS施工记录"
        if "雷达" in name:
            return "施工记录", "雷达工程"
        if "航向" in name:
            return "施工记录", "航向台施工记录"
        return "施工记录", "空管工程"

    if professional == "03_助航设施":
        if "灯光" in name:
            return "施工记录", "助航灯光工程"
        if "灯箱" in name:
            return "施工记录", "灯箱工程"
        if "标记牌" in name:
            return "施工记录", "标记牌工程"
        if "调光器" in name:
            return "施工记录", "调光器工程"
        if "papi" in lower:
            return "施工记录", "PAPI工程"
        return "施工记录", "助航设施工程"

    if professional == "04_弱电系统":
        if "监控" in name:
            return "施工记录", "监控系统"
        if "安防" in name:
            return "施工记录", "安防系统"
        if "网络" in name:
            return "施工记录", "网络系统"
        if "信息集成" in name:
            return "施工记录", "信息集成系统"
        return "施工记录", "弱电系统工程"

    if professional == "05_供油工程":
        if any(kw in name for kw in ["储油", "油罐"]):
            return "施工记录", "储油工程"
        if "管线" in name:
            return "施工记录", "管线工程"
        if "加油" in name:
            return "施工记录", "加油工程"
        return "施工记录", "供油工程"

    return "其他", "其他资料"


# ========== v7.2 C2: 图纸角色解耦 ==========
# is_drawing 标签：判断文件是否为图纸类
_DRAWING_KEYWORDS = [
    "平面图", "布置图", "设计图", "竣工图", "施工图", "剖面图",
    "断面图", "详图", "图纸", "示意图", "总平面", "管线图", "结构图",
]


def detect_is_drawing(filename: str) -> bool:
    """判断文件名是否表明这是一张图纸。"""
    name = filename.lower()
    return any(kw in name for kw in _DRAWING_KEYWORDS)


def detect_drawing_type(filename: str) -> Optional[str]:
    """v7.2 C2: 细分图纸类型（由文件名模式推断，不作硬判据）。

    返回:
        - "as_built"       竣工图
        - "process_sketch" 示意图/方案图
        - "design_basis"   设计依据图（施工图/平面图/布置图等）
        - None             非图纸
    """
    if not detect_is_drawing(filename):
        return None
    name = filename.lower()
    if "竣工图" in name or ("竣工" in name and "图" in name):
        return "as_built"
    if any(kw in name for kw in ["示意", "方案"]):
        return "process_sketch"
    return "design_basis"


def infer_doc_role(is_drawing: bool, stage: str, classification: str, drawing_type: Optional[str] = None) -> str:
    """根据 is_drawing + 项目阶段 + 图纸类型 + 分类，推断文档角色。
    返回: "reference" | "audited" | "general"
    - 竣工阶段图纸 → audited（需审核）
    - 示意图/方案图 → reference（过程示意，不审）
    - 施工阶段设计依据图 → reference
    - 非图纸 → 跟随 classification
    """
    if is_drawing:
        if drawing_type == "process_sketch":
            return "reference"
        # 竣工阶段（含"竣工""移交"等关键词）→ 图纸需审核
        if any(kw in stage for kw in ["竣工", "移交", "验收移交"]):
            return "audited"
        # 其他阶段（施工/分部分项验收）→ 图纸作为依据
        return "reference"
    # 非图纸：跟随原有 classification
    if classification == "reference_files":
        return "reference"
    if classification == "audited_files":
        return "audited"
    return "general"


# ========== OCR / PDF 提取 ==========
def call_ocr_image(
    file_path: Path,
    engine: str,
    text_out: Path,
    json_out: Path,
    preprocess: Optional[str] = None,
) -> Dict[str, Any]:
    """
    调用 OCR 对扫描件 PDF / 图片做 OCR。
    优先使用直接导入调用（避免子进程内存崩溃），降级为子进程。
    返回 {"text", "engine", "confidence", "items"}，失败时 confidence 为 0。
    """
    result: Dict[str, Any] = {"text": "", "engine": engine, "confidence": 0.0, "items": []}

    # 优先使用直接调用
    if _OCR_DIRECT_AVAILABLE and file_path.suffix.lower() == ".pdf":
        print(f"  [i] OCR 直接调用: engine={engine}, preprocess={preprocess or 'default'}", file=sys.stderr)
        try:
            ocr_result = _ocr_pdf_direct(
                str(file_path),
                dpi=72,
                engine=engine,
            )
            result["text"] = ocr_result.get("text", "")
            result["engine"] = ocr_result.get("engine", engine)
            result["confidence"] = ocr_result.get("confidence", 0.0)
            result["items"] = ocr_result.get("items", [])

            # 保存输出文件
            text_out.write_text(result["text"], encoding="utf-8")
            json_out.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            print(f"✅ OCR 完成，输出 {len(result['text'])} 字符", file=sys.stderr)
            return result
        except Exception as e:
            print(f"  [!] OCR 直接调用失败: {e}，降级为子进程", file=sys.stderr)
            import traceback
            traceback.print_exc()

    # 降级：子进程方式
    print(f"  [i] 调用 OCR: engine={engine}, preprocess={preprocess or 'default'}", file=sys.stderr)
    cmd = [
        sys.executable, str(OCR_IMAGE_PY),
        str(file_path),
        "--engine", engine,
        "--out", str(text_out),
        "--json-out", str(json_out),
        "--dpi", "72",
    ]
    if preprocess:
        cmd.extend(["--preprocess", preprocess])

    rc, stdout, stderr = run_subprocess(cmd)
    if stderr:
        print(stderr, file=sys.stderr)

    if rc == 0 and json_out.exists():
        try:
            raw = json.loads(json_out.read_text(encoding="utf-8"))
            result["text"] = raw.get("text", "")
            result["engine"] = raw.get("engine", engine)
            result["confidence"] = raw.get("confidence", 0.0)
            result["items"] = raw.get("items", [])
        except Exception as e:
            print(f"  [!] 解析 OCR JSON 失败: {e}", file=sys.stderr)
    else:
        print(f"  [!] OCR 子进程退出码: {rc}", file=sys.stderr)

    return result


def call_extract_pdf(file_path: Path, text_out: Path) -> Dict[str, Any]:
    """
    调用 extract_pdf.py 提取电子档 PDF 文字。
    返回 {"text", "engine", "confidence", "items": []}。
    """
    cmd = [sys.executable, str(EXTRACT_PDF_PY), str(file_path), "--out", str(text_out)]
    print("  [i] 调用 PyMuPDF 提取电子档 PDF", file=sys.stderr)
    rc, stdout, stderr = run_subprocess(cmd)
    if stderr:
        print(stderr, file=sys.stderr)

    text = ""
    if text_out.exists():
        text = text_out.read_text(encoding="utf-8")
    elif rc == 0:
        text = stdout

    return {
        "text": text,
        "engine": "PyMuPDF",
        "confidence": 1.0,
        "items": [],
    }


def call_extract_text(file_path: Path, text_out: Path) -> Dict[str, Any]:
    """
    直接读取文本文件（.txt/.md 等）内容。
    返回 {"text", "engine", "confidence", "items": []}。
    """
    print("  [i] 直接读取文本文件", file=sys.stderr)
    try:
        text = file_path.read_text(encoding="utf-8")
    except Exception as e:
        print(f"  [!] 读取文本文件失败: {e}", file=sys.stderr)
        text = ""
    text_out.write_text(text, encoding="utf-8")
    return {
        "text": text,
        "engine": "text",
        "confidence": 1.0,
        "items": [],
    }


def build_page_map(file_path: Path, is_scanned: bool, ocr_text: str, method: str) -> List[Dict[str, Any]]:
    """Build per-page text + image reference map."""
    page_map = []
    if method == "pymupdf" and HAS_PYMUPDF:
        try:
            import fitz
            doc = fitz.open(str(file_path))
            for i in range(len(doc)):
                page_text = doc[i].get_text("text").strip()
                images = doc[i].get_images()
                image_ref = None
                if images:
                    # Save page screenshot for PDFs with images
                    img_dir = file_path.parent / "_images"
                    img_dir.mkdir(exist_ok=True)
                    img_path = img_dir / f"{file_path.stem}_p{i+1}.png"
                    if not img_path.exists():
                        pix = doc[i].get_pixmap(dpi=150)
                        pix.save(str(img_path))
                    image_ref = str(img_path.name)
                page_map.append({"page": i + 1, "text": page_text, "image_ref": image_ref})
            doc.close()
        except Exception as e:
            page_map = [{"page": 1, "text": ocr_text, "image_ref": None}]
    else:
        page_map = [{"page": 1, "text": ocr_text, "image_ref": None}]
    return page_map


# ========== 结构化 rows 转换 ==========
def coerce_pile_value(field: str, raw: str) -> Any:
    """将字符串转换为字段对应的数据类型。"""
    raw = raw.strip()
    if field == "pile_no":
        return raw
    if field in ("start_time", "end_time", "remark"):
        return raw
    # 时间字符串保持原样
    if field in ("start_time", "end_time") and re.match(r"^\d{1,2}:\d{2}$", raw):
        return raw
    # 数字字段
    if raw == "":
        return None
    try:
        if "." in raw:
            return float(raw)
        return int(raw)
    except ValueError:
        return raw


# v7.2 C5: 模块级变量记录最近一次表头识别的原始 tokens（供 extract_header_info 使用）
_LAST_HEADER_RAW_TOKENS: Optional[List[str]] = None

# v7.2 C5: OCR 常见字符混淆归一化（路1 增强——OCR 认错表头字也能命中别名）
# 仅在表头匹配时使用，不改动原始 token
_OCR_HEADER_NORMALIZERS = [
    ("程高", "高程"),      # 高程 → 程高（常见 OCR 误读）
    ("实长", "实长"),      # 占位，保持列表非空可扩展
]


def _normalize_header_token(tok: str) -> str:
    """对表头 token 做 OCR 混淆归一化，返回归一化后的字符串。"""
    norm = tok
    for a, b in _OCR_HEADER_NORMALIZERS:
        norm = norm.replace(a, b)
    return norm


def _tokenize_table_line(line: str) -> List[str]:
    """表头行/数据行统一分词：优先多空格/制表符分隔，回退单空格。"""
    tokens = re.split(r"[\t\s]{2,}", line.strip())
    if len(tokens) < 3:
        tokens = line.strip().split()
    return tokens


def detect_header(line: str) -> Optional[Dict[int, str]]:
    """v7.2 C5: 表头检测——三路融合之第一路（关键词匹配 + OCR 归一化）。

    从 FIELD_ALIAS_MAP 聚合的槽位 schema 运行时获取关键词（修复原空字典 bug）。
    token 先做 OCR 混淆归一化（如"桩顶程高"→"桩顶高程"）再与别名比对。
    返回列索引 → 字段名映射；同时把原始 tokens 写入模块级变量供事后验证使用。
    """
    global _LAST_HEADER_RAW_TOKENS
    keywords = _get_pile_header_keywords()  # 运行时聚合，替代空字典 PILE_HEADER_KEYWORDS
    tokens = _tokenize_table_line(line)
    _LAST_HEADER_RAW_TOKENS = tokens
    mapping: Dict[int, str] = {}
    for idx, tok in enumerate(tokens):
        norm_tok = _normalize_header_token(tok)
        for field, kws in keywords.items():
            if any(kw in norm_tok for kw in kws):
                if idx not in mapping:
                    mapping[idx] = field
                break
    return mapping if len(mapping) >= 4 else None


def _validate_header_by_column_features(
    rows: List[Dict[str, Any]],
    header_map: Dict[int, str],
) -> Tuple[bool, float, Dict[str, str]]:
    """v7.2 C5: 三路融合之第二路——列特征验证。

    根据每列数据类型/范围验证 header_map 是否合理：
    - 数值列（design_length/diameter/actual_length 等）应为数字
    - 时间列（start_time/end_time）应为时间格式或可识别字符串
    - 桩号列（pile_no）应为 Z/D 开头或纯数字

    返回: (是否通过, 置信度[0-1], 各列问题说明)
    """
    schema = get_pile_slot_schema()
    numeric_slots = {s for s, d in schema.items() if d["data_type"] == "number"}
    time_slots = {s for s, d in schema.items() if d["data_type"] == "time"}

    issues: Dict[str, str] = {}
    if not rows:
        return True, 0.5, issues  # 无数据行无法验证，给中等置信度

    sample_size = min(len(rows), 10)
    sample = rows[:sample_size]
    total_checks = 0
    passed_checks = 0

    for idx, field in header_map.items():
        values = [r.get(field) for r in sample if r.get(field) is not None]
        if not values:
            continue
        total_checks += 1

        if field in numeric_slots:
            num_count = sum(1 for v in values if isinstance(v, (int, float)))
            if num_count / len(values) >= 0.7:
                passed_checks += 1
            else:
                issues[field] = f"数值列 {field} 仅 {num_count}/{len(values)} 为数字"
        elif field in time_slots:
            time_count = sum(
                1 for v in values
                if isinstance(v, str) and re.match(r"^\d{1,2}:\d{2}", v.strip())
            )
            if time_count / len(values) >= 0.6:
                passed_checks += 1
            else:
                issues[field] = f"时间列 {field} 仅 {time_count}/{len(values)} 为时间格式"
        elif field == "pile_no":
            pile_count = sum(
                1 for v in values
                if isinstance(v, str) and re.match(r"^[ZzDd]?\d", v.strip())
            )
            if pile_count / len(values) >= 0.7:
                passed_checks += 1
            else:
                issues[field] = f"桩号列格式异常"
        else:
            passed_checks += 1  # 非数值/时间/桩号列不验证

    if total_checks == 0:
        return True, 0.5, issues
    confidence = passed_checks / total_checks
    return confidence >= 0.7, confidence, issues


def _validate_header_by_math_chain(
    rows: List[Dict[str, Any]],
    header_map: Dict[int, str],
) -> Tuple[bool, float, List[str]]:
    """v7.2 C5: 三路融合之第三路——数学链约束验证。

    核心约束：实长 = 桩顶高程 − 桩底高程（容差 ±0.1m）
    若 header_map 中同时存在这三列，验证约束是否成立。
    返回: (是否通过, 置信度[0-1], 失败样本说明列表)
    """
    field_set = set(header_map.values())
    required = {"actual_length", "top_elev", "bottom_elev"}
    if not required.issubset(field_set):
        return True, 0.6, []  # 缺少必要列，无法验证，给中等置信度不阻断

    sample = [r for r in rows[:20] if all(r.get(f) is not None for f in required)]
    if len(sample) < 3:
        return True, 0.6, []  # 样本不足

    failures: List[str] = []
    passed = 0
    for r in sample:
        try:
            actual = float(r["actual_length"])
            top = float(r["top_elev"])
            bottom = float(r["bottom_elev"])
            computed = top - bottom
            if abs(actual - computed) <= 0.1 + 1e-9:  # 加 epsilon 避免浮点误差
                passed += 1
            else:
                failures.append(
                    f"桩号 {r.get('pile_no','?')}: 实长={actual}, 桩顶-桩底={computed:.2f}"
                )
        except (ValueError, TypeError):
            continue

    total = passed + len(failures)
    if total == 0:
        return True, 0.6, []
    confidence = passed / total
    return confidence >= 0.7, confidence, failures[:3]  # 最多记录3条样本


def _to_float(v: Any) -> Optional[float]:
    """尽力把值转成 float，失败返回 None。"""
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().rstrip("mM")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _infer_missing_slots_by_math_chain(
    raw_rows: List[List[str]],
    header_map: Dict[int, str],
) -> Tuple[Dict[int, str], List[str]]:
    """v7.2 C5: 数学链推断（路3 增强）——缺失三槽位时从未映射列推断。

    约束：实长 = 桩顶高程 − 桩底高程（容差 ±0.1m）
    场景：
      - 缺 actual_length，但 top/bottom 已映射 → target = mean(顶 - 底)
      - 缺 top_elev → target = mean(底 + 实长)
      - 缺 bottom_elev → target = mean(顶 - 实长)
      - 缺 top 和 bottom，但 actual_length 已映射 → 找两列差 ≈ 实长
    推断成立条件：样本 ≥3 行，匹配率 ≥80%，且列均值与 target 偏差 ≤0.1。

    返回: (补充后的 header_map, 推断说明列表)
    """
    if not raw_rows or len(raw_rows) < 3:
        return header_map, []

    mapped_fields = set(header_map.values())
    missing = {"top_elev", "bottom_elev", "actual_length"} - mapped_fields
    if not missing:
        return header_map, []

    # 提取每列数值（只取数字列）
    col_vals: Dict[int, List[float]] = {}
    for row in raw_rows:
        for idx, tok in enumerate(row):
            v = _to_float(tok)
            if v is not None:
                col_vals.setdefault(idx, []).append(v)

    # 未映射列范围取「已映射最大列 + 1」与「数据行最大列数」的较大者，
    # 否则当已映射列索引非连续（如仅映射列1/2）时，实长列（列3+）会被排除
    max_col = max((len(r) for r in raw_rows), default=0)
    unmapped_cols = [
        i for i in range(max(max(list(header_map.keys()) + [0]) + 1, max_col))
        if i not in header_map and len(col_vals.get(i, [])) >= 3
    ]
    if not unmapped_cols:
        return header_map, []

    def _find_col_approx(target_vals: List[float]) -> Optional[int]:
        """在未映射列中找均值最接近 target_vals 均值且逐行匹配率≥80% 的列。"""
        best_col = None
        best_err = None
        for c in unmapped_cols:
            vals = col_vals[c]
            if len(vals) < 3:
                continue
            n = min(len(vals), len(target_vals))
            if n < 3:
                continue
            # 逐行校验（前 n 行对齐）
            matches = sum(
                1 for k in range(n)
                if abs(vals[k] - target_vals[k]) <= 0.1 + 1e-9
            )
            if matches / n < 0.8:
                continue
            mean_err = abs(sum(vals[:n]) / n - sum(target_vals[:n]) / n)
            if best_err is None or mean_err < best_err:
                best_err = mean_err
                best_col = c
        return best_col

    inferred: Dict[int, str] = {}
    notes: List[str] = []

    # 每行目标值生成器
    def _row_target() -> List[float]:
        """根据缺失集合生成 target 序列（按行）。"""
        targets: List[float] = []
        for row in raw_rows:
            vals: Dict[str, float] = {}
            for idx, field in header_map.items():
                v = _to_float(row[idx]) if idx < len(row) else None
                if v is not None:
                    vals[field] = v
            if "top_elev" in vals and "bottom_elev" in vals:
                # 顶/底都有：可推实长
                if "actual_length" in missing:
                    targets.append(vals["top_elev"] - vals["bottom_elev"])
            if "bottom_elev" in vals and "actual_length" in vals:
                if "top_elev" in missing:
                    targets.append(vals["bottom_elev"] + vals["actual_length"])
            if "top_elev" in vals and "actual_length" in vals:
                if "bottom_elev" in missing:
                    targets.append(vals["top_elev"] - vals["actual_length"])
        return targets

    # 场景1：缺 top 和 bottom，但 actual_length 已映射 → 找两列差 ≈ 实长
    if "top_elev" in missing and "bottom_elev" in missing and "actual_length" in mapped_fields:
        target = [
            _to_float(row[col_idx])
            for col_idx, field in header_map.items()
            if field == "actual_length"
            for row in raw_rows
        ]
        if len(target) >= 3:
            best_pair = None
            best_pair_err = None
            for i in unmapped_cols:
                for j in unmapped_cols:
                    if i == j:
                        continue
                    vals_i = col_vals[i]
                    vals_j = col_vals[j]
                    n = min(len(vals_i), len(vals_j), len(target))
                    if n < 3:
                        continue
                    matches = sum(
                        1 for k in range(n)
                        if vals_i[k] is not None and vals_j[k] is not None and target[k] is not None
                        and abs((vals_i[k] - vals_j[k]) - target[k]) <= 0.1 + 1e-9
                    )
                    if matches / n < 0.8:
                        continue
                    # 过滤 None 计算均值
                    valid_vals_i = [v for v in vals_i[:n] if v is not None]
                    valid_vals_j = [v for v in vals_j[:n] if v is not None]
                    valid_target = [v for v in target[:n] if v is not None]
                    nv = min(len(valid_vals_i), len(valid_vals_j), len(valid_target))
                    if nv < 3:
                        continue
                    mean_err = abs(
                        (sum(valid_vals_i[:nv]) - sum(valid_vals_j[:nv])) / nv
                        - sum(valid_target[:nv]) / nv
                    )
                    if best_pair_err is None or mean_err < best_pair_err:
                        best_pair_err = mean_err
                        best_pair = (i, j)
            if best_pair:
                hi, lo = best_pair
                # 高者通常数值更大（桩顶高程 > 桩底高程）
                if sum(col_vals[hi]) < sum(col_vals[lo]):
                    hi, lo = lo, hi
                inferred[hi] = "top_elev"
                inferred[lo] = "bottom_elev"
                notes.append(f"数学链推断: 列{hi}→桩顶高程, 列{lo}→桩底高程（顶-底≈实长）")

    # 场景2：单一缺失槽位
    elif len(missing) == 1:
        targets = _row_target()
        if len(targets) >= 3:
            col = _find_col_approx(targets)
            if col is not None:
                slot = next(iter(missing))
                inferred[col] = slot
                notes.append(f"数学链推断: 列{col}→{slot}（满足实长=顶-底约束）")

    if not inferred:
        return header_map, notes

    # 合并：已映射列优先，推断列不覆盖已映射
    merged = dict(header_map)
    for idx, field in inferred.items():
        if idx not in merged:
            merged[idx] = field
    return merged, notes


def extract_header_info(text: str, doc_type: str) -> Dict[str, Any]:
    """v7.2 C5: 综合表头识别——三路融合 + 数学链推断 + 人工确认入口。

    返回结构：
        {
            "raw_headers": [...],          # 原始表头 tokens
            "header_mapping": {...},       # 列索引(字符串) → 字段名（含推断补全）
            "header_source": "keyword|keyword_math|none",
            "header_confidence": 0.0~1.0,
            "math_chain_verified": bool,
            "math_chain_failures": [...],
            "math_chain_inferred": [...],  # 数学链推断说明
            "column_feature_issues": {...},
            "needs_human_confirm": bool,   # 置信度<0.7 时为 True
        }
    """
    global _LAST_HEADER_RAW_TOKENS
    lower = doc_type.lower()
    is_pile = any(kw in lower for kw in ["碎石桩", "cfg", "桩"])
    if not is_pile:
        # v7.2 通用表格：检测表头并记录，支持 data-editor 表头映射 Tab 展示与人工核对
        generic = detect_generic_header_from_text(text)
        if generic:
            conf = generic["header_confidence"]
            return {
                "raw_headers": generic["raw_headers"],
                "header_mapping": generic["header_mapping"],
                "header_source": "generic",
                "header_confidence": conf,
                "math_chain_verified": True,   # 通用表格不适用数学链
                "math_chain_failures": [],
                "math_chain_inferred": [],
                "column_feature_issues": {},
                "needs_human_confirm": conf < 0.7,
            }
        return {
            "raw_headers": [],
            "header_mapping": {},
            "header_source": "none",
            "header_confidence": 1.0,
            "math_chain_verified": True,
            "math_chain_failures": [],
            "math_chain_inferred": [],
            "column_feature_issues": {},
            "needs_human_confirm": False,
        }

    # 找到表头行，同时收集原始 token 行供数学链推断
    header_map: Optional[Dict[int, str]] = None
    raw_tokens: List[str] = []
    rows_for_validation: List[Dict[str, Any]] = []
    raw_rows: List[List[str]] = []  # 原始 token 行（含未映射列）
    page = 1
    line_no = 1
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = re.match(r"===\s*第\s*(\d+)\s*页\s*===", line)
        if m:
            page = int(m.group(1))
            header_map = None
            continue
        if header_map is None:
            header_map = detect_header(line)
            if header_map:
                raw_tokens = list(_LAST_HEADER_RAW_TOKENS or [])
                continue
        if header_map:
            tokens = _tokenize_table_line(line)
            record = parse_pile_data_line(line, header_map, page, line_no)
            if record:
                rows_for_validation.append(record)
                raw_rows.append(tokens)
                line_no += 1
                if len(rows_for_validation) >= 30:
                    break

    if not header_map:
        return {
            "raw_headers": [],
            "header_mapping": {},
            "header_source": "none",
            "header_confidence": 0.0,
            "math_chain_verified": False,
            "math_chain_failures": [],
            "math_chain_inferred": [],
            "column_feature_issues": {},
            "needs_human_confirm": True,
        }

    # 路3增强：数学链推断补全缺失三槽位（OCR 表头认错时仍能对上）
    header_map, math_notes = _infer_missing_slots_by_math_chain(raw_rows, header_map)
    source = "keyword_math" if math_notes else "keyword"

    # 第二路：列特征验证
    feat_ok, feat_conf, feat_issues = _validate_header_by_column_features(
        rows_for_validation, header_map
    )
    # 第三路：数学链约束验证
    math_ok, math_conf, math_failures = _validate_header_by_math_chain(
        rows_for_validation, header_map
    )

    # 综合置信度：关键词匹配(0.5) + 列特征(0.3) + 数学链(0.2)
    # 有数学链推断时置信度不再因缺列而降级，且推断本身是强信号
    if math_notes:
        confidence = 0.5 + 0.3 * feat_conf + 0.2 * max(math_conf, 0.7)
    else:
        confidence = 0.5 + 0.3 * feat_conf + 0.2 * math_conf
    confidence = min(confidence, 1.0)

    needs_confirm = confidence < 0.7

    return {
        "raw_headers": raw_tokens,
        "header_mapping": {str(k): v for k, v in header_map.items()},
        "header_source": source,
        "header_confidence": round(confidence, 3),
        "math_chain_verified": math_ok,
        "math_chain_failures": math_failures,
        "math_chain_inferred": math_notes,
        "column_feature_issues": feat_issues,
        "needs_human_confirm": needs_confirm,
    }


def parse_pile_data_line(line: str, header_map: Dict[int, str], page: int, line_no: int) -> Optional[Dict[str, Any]]:
    """按表头映射解析一行碎石桩/桩基数据。"""
    tokens = re.split(r"[\t\s]{2,}", line.strip())
    if len(tokens) < len(header_map):
        tokens += [""] * (len(header_map) - len(tokens))

    record: Dict[str, Any] = {"page": page, "line_no": line_no}
    for idx, field in header_map.items():
        if idx < len(tokens):
            record[field] = coerce_pile_value(field, tokens[idx])

    # 至少要有桩号列才认为是数据行
    if not record.get("pile_no"):
        return None
    return record


def heuristic_pile_row(line: str, page: int, line_no: int) -> Optional[Dict[str, Any]]:
    """无表头时的启发式解析：按数值范围粗略匹配碎石桩字段。"""
    tokens = line.strip().split()
    if len(tokens) < 6:
        return None

    # 找桩号：Z 开头或 2 开头
    pile_idx = None
    for i, t in enumerate(tokens):
        if re.match(r"^[Zz2][\.\,\;\:\-]?[4-9]\d{1,2}[A-Da-d]?$", t.strip()):
            pile_idx = i
            break
    if pile_idx is None:
        return None

    nums = []
    for t in tokens[pile_idx + 1:]:
        s = t.strip()
        if re.match(r"^[\d\.]+$", s):
            try:
                nums.append(float(s))
            except ValueError:
                pass
        elif re.match(r"^\d{1,2}:\d{2}$", s):
            nums.append(s)

    record: Dict[str, Any] = {"page": page, "line_no": line_no, "pile_no": tokens[pile_idx].strip()}
    field_order = [
        "design_length", "diameter", "bottom_elev", "top_elev", "actual_length",
        "current", "re_penetration", "volume", "filling_coeff", "verticality",
    ]
    num_idx = 0
    for field in field_order:
        if num_idx >= len(nums):
            break
        v = nums[num_idx]
        if isinstance(v, str):
            # 时间字段只出现在最后两个位置
            if field in ("start_time", "end_time"):
                record[field] = v
            num_idx += 1
            continue
        record[field] = v
        num_idx += 1

    # 最后两个数字可能是开始/结束时间字符串
    times = [v for v in nums if isinstance(v, str)]
    if "start_time" not in record and len(times) >= 1:
        record["start_time"] = times[0]
    if "end_time" not in record and len(times) >= 2:
        record["end_time"] = times[1]

    return record


def parse_pile_rows(text: str, start_line_no: int = 1) -> List[Dict[str, Any]]:
    """从 OCR/PDF 文本中解析碎石桩类表格 rows。"""
    rows: List[Dict[str, Any]] = []
    page = 1
    line_no = start_line_no
    header_map: Optional[Dict[int, str]] = None

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        # 页分隔符
        m = re.match(r"===\s*第\s*(\d+)\s*页\s*===", line)
        if m:
            page = int(m.group(1))
            header_map = None
            continue

        # 表头检测
        if header_map is None:
            header_map = detect_header(line)
            if header_map:
                continue

        if header_map:
            record = parse_pile_data_line(line, header_map, page, line_no)
        else:
            record = heuristic_pile_row(line, page, line_no)

        if record:
            rows.append(record)
            line_no += 1

    return rows


def parse_generic_rows(text: str, start_line_no: int = 1) -> List[Dict[str, Any]]:
    """从文本中提取通用 rows：page / line_no / raw_text。"""
    rows: List[Dict[str, Any]] = []
    page = 1
    line_no = start_line_no
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = re.match(r"===\s*第\s*(\d+)\s*页\s*===", line)
        if m:
            page = int(m.group(1))
            continue
        rows.append({"page": page, "line_no": line_no, "raw_text": line})
        line_no += 1
    return rows


# ========== v7.2 通用表格提取（非桩基类文档结构化）==========
# 通用表头关键词加成词表（命中越多越可能是表头行）
_GENERIC_HEADER_KEYWORDS = {
    "编号", "序号", "项目", "名称", "规格", "型号", "单位", "数量", "日期",
    "时间", "桩号", "高程", "直径", "实长", "灌入量", "充盈系数", "垂直度",
    "备注", "签字", "施工", "监理", "检查", "允许偏差", "实测", "检验",
    "结果", "设计", "实际", "标准", "规定", "依据", "部位", "里程", "标段",
    "项目名称", "分项工程", "分部工程", "施工单位", "建设", "设计值",
    "允许值", "质量情况", "评定", "等级", "结论", "频率", "批次",
}


def _is_text_token(tok: str) -> bool:
    """判断 token 是否为文本型（表头候选）。

    判定规则：
      - 含中文 → 文本（表头主体）
      - 纯字母（无数字）→ 文本（如 Item/Date/No）
      - 字母+数字混合（如 Z420/H-001）→ 非文本（代码/标识符，属数据值）
      - 纯数字/日期/时间 → 非文本
    """
    s = tok.strip()
    if not s:
        return False
    # 纯数字（含小数、千分位）
    if re.match(r"^[\d\.\,\-]+$", s):
        return False
    # 时间格式 HH:MM
    if re.match(r"^\d{1,2}:\d{2}$", s):
        return False
    # 日期格式
    if re.match(r"^\d{4}[-/年]\d{1,2}", s):
        return False
    # 含中文 → 文本
    if re.search(r"[\u4e00-\u9fa5]", s):
        return True
    # 纯字母（无数字）→ 文本
    if re.match(r"^[a-zA-Z]+$", s):
        return True
    # 字母+数字混合 → 非文本（代码/标识符）
    return False


def _normalize_col_name(tok: str) -> str:
    """归一化列名：去空白、去尾部冒号/星号/换行符。"""
    s = tok.strip()
    s = s.rstrip(":：*＊\n\r")
    return s.strip()


def _coerce_generic_value(raw: str) -> Any:
    """通用表格单元格值类型转换：数字→int/float，否则字符串。"""
    s = raw.strip()
    if s == "":
        return ""
    # 占位符保留为字符串
    if s in ("-", "/", "—", "--", "/"):
        return s
    # 纯整数
    if re.match(r"^-?\d+$", s):
        try:
            return int(s)
        except ValueError:
            return s
    # 小数
    if re.match(r"^-?\d+\.\d+$", s):
        try:
            return float(s)
        except ValueError:
            return s
    return s


def detect_generic_header(line: str) -> Optional[Tuple[Dict[int, str], List[str], float]]:
    """通用表格表头检测：识别文本主导的多列行作为表头。

    判定规则：
      1. 分词后 token 数 ≥ 3
      2. 文本型 token 占比 ≥ 50% 且文本型 token 数 ≥ 2
      3. 有效列名数 ≥ 3（去重后）

    返回: (col_idx → col_name 映射, 原始 tokens, 置信度) 或 None。
    """
    tokens = _tokenize_table_line(line)
    if len(tokens) < 3:
        return None

    text_tokens = [t for t in tokens if _is_text_token(t)]
    text_ratio = len(text_tokens) / len(tokens)
    if text_ratio < 0.5 or len(text_tokens) < 2:
        return None

    # 关键词加成
    kw_hits = sum(1 for t in tokens if any(kw in t for kw in _GENERIC_HEADER_KEYWORDS))
    confidence = 0.6 + 0.3 * min(text_ratio, 1.0) - 0.1
    if kw_hits >= 2:
        confidence = min(confidence + 0.15, 0.95)
    elif kw_hits >= 1:
        confidence = min(confidence + 0.08, 0.95)

    # 构建列映射（列名去重）
    seen_names: Dict[str, int] = {}
    header_map: Dict[int, str] = {}
    for idx, tok in enumerate(tokens):
        name = _normalize_col_name(tok)
        if not name:
            continue
        if name in seen_names:
            seen_names[name] += 1
            name = f"{name}_{seen_names[name]}"
        else:
            seen_names[name] = 1
        header_map[idx] = name

    if len(header_map) < 3:
        return None

    return header_map, tokens, round(confidence, 3)


def parse_generic_table(text: str, start_line_no: int = 1) -> List[Dict[str, Any]]:
    """从文本中提取通用表格结构化行。

    检测表头行 + 列边界，按表头列名提取结构化字段。
    适用于检验批、隐蔽工程记录、混凝土施工记录等非桩基类表格文档。

    流程：
      1. 逐行扫描，遇页分隔符翻页并重置表头
      2. 首个通过 detect_generic_header 的行作为表头
      3. 后续行按表头列映射解析为结构化记录
      4. 跳过空行和全空数据行
      5. 数据行 < 2 行视为非表格，返回空列表（由调用方回退到 parse_generic_rows）

    返回: 结构化行列表，每行含 row_index/page/line_no + 列名字段。
          未检测到表头或数据行不足时返回空列表。
    """
    rows: List[Dict[str, Any]] = []
    page = 1
    line_no = start_line_no
    row_index = 0
    header_map: Optional[Dict[int, str]] = None

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        # 页分隔符：翻页重置表头
        m = re.match(r"===\s*第\s*(\d+)\s*页\s*===", line)
        if m:
            page = int(m.group(1))
            header_map = None
            continue

        # 表头检测
        if header_map is None:
            detected = detect_generic_header(line)
            if detected:
                header_map, _tokens, _conf = detected
                continue
            # 未检测到表头则跳过（不把前置文本当数据）
            continue

        # 数据行解析
        tokens = _tokenize_table_line(line)
        if len(tokens) < 2:
            continue

        record: Dict[str, Any] = {
            "row_index": row_index + 1,
            "page": page,
            "line_no": line_no,
        }
        all_empty = True
        for idx, col_name in header_map.items():
            if idx < len(tokens):
                val = _coerce_generic_value(tokens[idx])
                record[col_name] = val
                if val not in (None, ""):
                    all_empty = False
            else:
                record[col_name] = None

        if all_empty:
            continue

        rows.append(record)
        row_index += 1
        line_no += 1

    # 数据行不足 2 行视为非表格，回退纯文本
    if len(rows) < 2:
        return []

    return rows


def detect_generic_header_from_text(text: str) -> Optional[Dict[str, Any]]:
    """扫描全文，返回首个通用表头信息。供 extract_header_info 使用。

    返回: {"raw_headers": [...], "header_mapping": {...}, "header_confidence": float} 或 None。
    """
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = re.match(r"===\s*第\s*(\d+)\s*页\s*===", line)
        if m:
            continue
        detected = detect_generic_header(line)
        if detected:
            header_map, tokens, confidence = detected
            return {
                "raw_headers": tokens,
                "header_mapping": {str(k): v for k, v in header_map.items()},
                "header_confidence": confidence,
            }
    return None


def build_rows(text: str, doc_type: str) -> List[Dict[str, Any]]:
    """根据 doc_type 选择解析策略。

    - 桩基类文档（碎石桩/CFG/桩）→ parse_pile_rows（表头三路融合 + 数学链约束）
    - 非桩基类文档 → 先尝试 parse_generic_table（通用表格提取）
      - 检测到表头且数据行≥2 → 返回结构化行
      - 否则回退 parse_generic_rows（纯文本行 page/line_no/raw_text）
    """
    lower = doc_type.lower()
    is_pile = any(kw in lower for kw in ["碎石桩", "cfg", "桩"])
    if is_pile:
        return parse_pile_rows(text)
    # 通用表格提取：先尝试结构化表格，失败回退纯文本行
    rows = parse_generic_table(text)
    if rows:
        return rows
    return parse_generic_rows(text)


# ========== 行数校验 ==========
def get_expected_rows(
    rel_path: Path,
    pages: int,
    file_type: str,
    expected_map: Dict[str, int],
) -> Optional[int]:
    """获取文件的期望行数。"""
    name = rel_path.name
    # 1. 显式 expected-rows 配置
    for pattern, n in expected_map.items():
        if pattern in name:
            return n
    # 2. 按页推断
    if pages and pages > 0:
        if file_type.upper() == "PDF":
            return pages * DEFAULT_ROWS_PER_PDF_PAGE
        return pages * DEFAULT_ROWS_PER_IMAGE_PAGE
    return None


def check_row_count(actual: int, expected: Optional[int]) -> Tuple[bool, Optional[float]]:
    """返回 (是否通过, 实际/预期比例)。"""
    if expected is None or expected <= 0:
        return True, None
    ratio = actual / expected
    return ratio >= ROW_COUNT_THRESHOLD, ratio


# ========== MD 预览 ==========
def write_md_preview(
    data_file: Path,
    md_file: Path,
    meta: Dict[str, Any],
) -> None:
    """根据 {name}.json 生成只读 Markdown 预览。"""
    data = load_json(data_file) or {}
    rows = data.get("rows", [])

    lines: List[str] = []
    lines.append(f"# {meta.get('doc_type', '文档')} - 数据预览")
    lines.append("")
    lines.append(f"- **原始文件**: `{meta.get('original_file', '')}`")
    lines.append(f"- **OCR 引擎**: {meta.get('ocr_engine', '')}")
    conf = meta.get("ocr_confidence", 0.0) or 0.0
    lines.append(f"- **平均置信度**: {conf:.1%}")
    lines.append(f"- **页数**: {meta.get('pages', '')}")
    lines.append(f"- **生成时间**: {meta.get('ocr_completed_at', '')}")
    lines.append(f"- **状态**: {meta.get('ocr_status', '')}")
    lines.append("")

    if not rows:
        lines.append("> 未识别到数据行")
    else:
        headers = list(rows[0].keys())
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for row in rows:
            vals = [str(v) if v is not None else "" for v in row.values()]
            lines.append("| " + " | ".join(vals) + " |")

    md_file.parent.mkdir(parents=True, exist_ok=True)
    md_file.write_text("\n".join(lines), encoding="utf-8")


# ========== 质量检测 / 混淆检测 ==========
def call_data_quality_check(data_file: Path) -> Tuple[Dict[str, Any], int]:
    """调用 data_quality_check.py，返回结果字典与告警总数。"""
    cmd = [sys.executable, str(DQ_CHECK_PY), str(data_file), "--pretty"]
    rc, stdout, stderr = run_subprocess(cmd)
    if stderr:
        print(stderr, file=sys.stderr)
    if rc != 0:
        print(f"  [!] 数据质量检测失败，退出码 {rc}", file=sys.stderr)
        return {"status": "error", "summary": {"total_warnings": 0}, "warnings": []}, 0
    try:
        result = json.loads(stdout)
        alerts = result.get("summary", {}).get("total_warnings", 0)
        return result, alerts
    except Exception as e:
        print(f"  [!] 解析数据质量结果失败: {e}", file=sys.stderr)
        return {"status": "error", "summary": {"total_warnings": 0}, "warnings": []}, 0


def call_ocr_confusion_check(data_file: Path) -> Tuple[Dict[str, Any], int]:
    """调用 ocr_confusion_check.py，返回结果字典与存疑总数。"""
    cmd = [sys.executable, str(CONFUSION_CHECK_PY), str(data_file), "--pretty"]
    rc, stdout, stderr = run_subprocess(cmd)
    if stderr:
        print(stderr, file=sys.stderr)
    if rc != 0:
        print(f"  [!] 混淆检测失败，退出码 {rc}", file=sys.stderr)
        return {"status": "error", "summary": {"total_suspects": 0}, "suspects": []}, 0
    try:
        result = json.loads(stdout)
        suspects = result.get("summary", {}).get("total_suspects", 0)
        return result, suspects
    except Exception as e:
        print(f"  [!] 解析混淆检测结果失败: {e}", file=sys.stderr)
        return {"status": "error", "summary": {"total_suspects": 0}, "suspects": []}, 0


# ========== index.json ==========
def ensure_index(
    project_path: Path,
    out_base: Path,
    preconditions: Dict[str, Any],
) -> Dict[str, Any]:
    """读取或初始化 index.json。"""
    index_path = out_base / "index.json"
    existing = load_json(index_path)
    if existing and existing.get("schema_version") == "1.0":
        existing.setdefault("preconditions", preconditions)
        return existing

    return {
        "schema_version": "1.0",
        "project_name": project_path.name,
        "project_path": str(project_path.resolve()),
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "stage": "foundation_built",
        "preconditions": preconditions,
        "file_classification": {
            "audited_files": [],
            "reference_files": [],
            "excluded_files": [],
        },
        "documents": [],
        "corrections": {"total": 0, "files": []},
        "gaps": [],
        "audit_logs": [],
    }


def next_doc_id(index: Dict[str, Any]) -> str:
    """生成下一个 DOC-xxx 编号。"""
    existing = [d.get("id", "") for d in index.get("documents", [])]
    nums = []
    for e in existing:
        m = re.match(r"DOC-(\d+)", e)
        if m:
            nums.append(int(m.group(1)))
    nxt = max(nums, default=0) + 1
    return f"DOC-{nxt:03d}"


def find_doc_by_original(index: Dict[str, Any], original: str) -> Optional[Dict[str, Any]]:
    """根据 original_file 查找已有文档记录。"""
    for d in index.get("documents", []):
        if d.get("original_file") == original:
            return d
    return None


def sync_documents_with_classification(index: Dict[str, Any]) -> None:
    """
    根据当前 file_classification 重新同步 documents 数组。
    - 保留 original_file 仍在 audited_files 中的条目；
    - 人工核对过的条目（human_verified=true）即使不再被审核也保留，但标记 stale；
    - 其余 stale 条目清理。
    """
    audited_files = set(index.get("file_classification", {}).get("audited_files", []))
    docs = index.setdefault("documents", [])
    kept: List[Dict[str, Any]] = []
    for doc in docs:
        original = doc.get("original_file")
        if original in audited_files:
            doc.pop("stale", None)
            kept.append(doc)
        elif doc.get("human_verified") is True:
            doc["stale"] = True
            kept.append(doc)
        # 其他情况直接丢弃
    index["documents"] = kept


def update_index_for_doc(
    index: Dict[str, Any],
    doc: Dict[str, Any],
) -> None:
    """用新的文档记录更新 index.json（新增或替换）。"""
    docs = index.setdefault("documents", [])
    for i, d in enumerate(docs):
        if d.get("id") == doc.get("id"):
            docs[i] = doc
            return
    docs.append(doc)


# ========== 模板复制 ==========
def copy_web_templates(project_path: Path) -> None:
    """将 Web 模板复制到项目文件夹根目录（含 PDF.js 离线依赖）。"""
    # 模板文件清单：(源文件名, 目标文件名, 描述)
    template_files = [
        ("data-editor.html", "data-editor.html", "数据编辑器"),
        ("project-dashboard.html", "项目总览.html", "项目总览仪表盘"),
        ("tokens.css", "tokens.css", "统一设计令牌"),
        ("pdf.min.js", "pdf.min.js", "PDF.js 主库（离线）"),
        ("pdf.worker.min.js", "pdf.worker.min.js", "PDF.js Worker（离线）"),
    ]

    for src_name, dst_name, desc in template_files:
        src = TEMPLATES_DIR / src_name
        dst = project_path / dst_name
        if src.exists():
            shutil.copy2(src, dst)
            print(f"  ✓ 复制模板: {dst}（{desc}）", file=sys.stderr)
        else:
            print(f"  [!] 模板缺失: {src}（{desc}）", file=sys.stderr)


# ========== 断档检测（N-09） ==========
def _parse_pile_no(raw: str) -> Optional[int]:
    """解析桩号字符串，提取数值部分。

    支持格式：
      - Z420 → 420
      - Z-420 → 420
      - PHC-001 → 1
      - K0+120 → None（里程桩号，不适合数值连续性检测）
      - 纯数字 → 数字本身
    """
    if not raw or not isinstance(raw, str):
        return None
    raw = raw.strip()
    # 里程桩号（Kxx+xxx 格式），不参与数值连续性检测
    if re.match(r'^K\d+[\+]\d+', raw, re.IGNORECASE):
        return None
    # 提取最后的连续数字
    m = re.search(r'(\d+)$', raw)
    if m:
        return int(m.group(1))
    return None


def _extract_field_from_rows(rows: List[Dict[str, Any]], field_names: List[str]) -> List[Any]:
    """从 rows 中提取指定字段的值列表（按行顺序，去空）。"""
    values = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        for fn in field_names:
            v = row.get(fn)
            if v is not None and v != "":
                values.append(v)
                break
    return values


def _detect_sequence_gaps(
    numbers: List[int],
    doc_label: str,
    professional: str,
    gap_type: str = "pile_no",
) -> List[Dict[str, Any]]:
    """检测数值序列中的断档。

    Args:
        numbers: 已排序的数值列表
        doc_label: 文档标识（用于描述）
        professional: 专业分类
        gap_type: 断档类型（pile_no / serial_no / date_seq）

    Returns:
        断档列表，每项包含 range_start, range_end, missing_count, description
    """
    if len(numbers) < 2:
        return []

    gaps = []
    # 先排序去重
    sorted_unique = sorted(set(numbers))
    min_val = sorted_unique[0]
    max_val = sorted_unique[-1]
    expected_range = max_val - min_val + 1
    actual_count = len(sorted_unique)

    if actual_count >= expected_range:
        return []  # 无断档

    # 找出缺失的具体数值
    missing = sorted(set(range(min_val, max_val + 1)) - set(sorted_unique))

    if not missing:
        return []

    # 将连续缺失合并为区间
    range_start = missing[0]
    range_end = missing[0]
    for m in missing[1:]:
        if m == range_end + 1:
            range_end = m
        else:
            gaps.append({
                "type": gap_type,
                "professional": professional,
                "range_start": str(range_start),
                "range_end": str(range_end) if range_end != range_start else str(range_start),
                "missing_count": range_end - range_start + 1,
                "description": f"{doc_label}：{gap_type} {range_start}-{range_end} 缺失（共 {range_end - range_start + 1} 个）",
                "detected_at": now_iso(),
            })
            range_start = m
            range_end = m

    # 最后一组
    gaps.append({
        "type": gap_type,
        "professional": professional,
        "range_start": str(range_start),
        "range_end": str(range_end) if range_end != range_start else str(range_start),
        "missing_count": range_end - range_start + 1,
        "description": f"{doc_label}：{gap_type} {range_start}-{range_end} 缺失（共 {range_end - range_start + 1} 个）",
        "detected_at": now_iso(),
    })

    return gaps


def detect_gaps(index: Dict[str, Any], out_base: Path) -> List[Dict[str, Any]]:
    """检测数据底座中的断档（桩号/日期/编号连续性）。

    从 index.json 的 documents 数组中读取所有已完成 OCR 的文件，
    加载其结构化数据，检测以下类型的断档：
    1. 桩号连续性：桩基施工记录中桩号是否连续
    2. 日期连续性：施工日志中日期是否连续
    3. 编号连续性：检验批/隐蔽工程编号是否连续

    Returns:
        断档列表，同时写入 index["gaps"]
    """
    all_gaps: List[Dict[str, Any]] = []

    # 桩号相关字段名（按优先级）
    PILE_NO_FIELDS = ["pile_no", "桩号", "序号/桩", "桩号/序号", "编号"]
    DATE_FIELDS = ["date", "日期", "施工日期", "记录日期", "start_time", "开始时间"]
    SERIAL_FIELDS = ["serial_no", "编号", "序号", "检验批编号", "no"]

    # 桩相关文档类型关键词
    PILE_DOC_KEYWORDS = ["桩", "碎石桩", "PHC", "CFG", "管桩", "灌注桩"]

    for doc in index.get("documents", []):
        doc_id = doc.get("id", "")
        doc_type = doc.get("doc_type", "")
        professional = doc.get("professional", "")
        data_file_rel = doc.get("data_file", "")
        ocr_status = doc.get("ocr_status", "")

        if ocr_status != "completed":
            continue

        # v7.2: Excel/元数据类文档没有 data_file（不参与断档检测）
        if not data_file_rel:
            continue

        data_path = out_base / data_file_rel
        if not data_path.exists():
            continue

        data = load_json(data_path)
        if not data:
            continue

        rows = data.get("rows", [])
        if not rows:
            continue

        doc_label = f"{doc_type}（{doc.get('original_file', doc_id)}）"

        # === 1. 桩号连续性检测 ===
        is_pile_doc = any(kw in doc_type for kw in PILE_DOC_KEYWORDS)
        if is_pile_doc:
            pile_raw = _extract_field_from_rows(rows, PILE_NO_FIELDS)
            pile_nums = []
            for p in pile_raw:
                num = _parse_pile_no(str(p))
                if num is not None:
                    pile_nums.append(num)

            if pile_nums:
                pile_nums.sort()
                pile_gaps = _detect_sequence_gaps(
                    pile_nums, doc_label, professional, gap_type="pile_no"
                )
                all_gaps.extend(pile_gaps)
                if pile_gaps:
                    print(f"  🔍 桩号断档: {doc_label} — 发现 {len(pile_gaps)} 处断档", file=sys.stderr)

        # === 2. 日期连续性检测 ===
        dates_raw = _extract_field_from_rows(rows, DATE_FIELDS)
        if dates_raw:
            date_nums = []
            for d in dates_raw:
                num = _parse_date_to_ordinal(str(d))
                if num is not None:
                    date_nums.append(num)
            if len(date_nums) >= 2:
                date_nums.sort()
                date_gaps = _detect_sequence_gaps(
                    date_nums, doc_label, professional, gap_type="date"
                )
                # 日期断档需要转换为可读格式
                for g in date_gaps:
                    try:
                        g["range_start"] = _ordinal_to_date_str(int(g["range_start"]))
                        g["range_end"] = _ordinal_to_date_str(int(g["range_end"]))
                        g["description"] = (
                            f"{doc_label}：日期 {g['range_start']} 至 {g['range_end']} "
                            f"缺失（共 {g['missing_count']} 天）"
                        )
                    except (ValueError, TypeError):
                        pass
                all_gaps.extend(date_gaps)
                if date_gaps:
                    print(f"  🔍 日期断档: {doc_label} — 发现 {len(date_gaps)} 处断档", file=sys.stderr)

        # === 3. 编号连续性检测（非桩类文档） ===
        if not is_pile_doc:
            serial_raw = _extract_field_from_rows(rows, SERIAL_FIELDS)
            serial_nums = []
            for s in serial_raw:
                num = _parse_pile_no(str(s))  # 复用同一解析逻辑
                if num is not None:
                    serial_nums.append(num)

            if len(serial_nums) >= 3:
                serial_nums.sort()
                serial_gaps = _detect_sequence_gaps(
                    serial_nums, doc_label, professional, gap_type="serial_no"
                )
                all_gaps.extend(serial_gaps)
                if serial_gaps:
                    print(f"  🔍 编号断档: {doc_label} — 发现 {len(serial_gaps)} 处断档", file=sys.stderr)

    # 写入 index
    index["gaps"] = all_gaps

    if all_gaps:
        print(f"\n⚠️  断档检测完成：共发现 {len(all_gaps)} 处断档", file=sys.stderr)
    else:
        print(f"\n✅ 断档检测完成：未发现断档", file=sys.stderr)

    return all_gaps


def _parse_date_to_ordinal(date_str: str) -> Optional[int]:
    """将日期字符串转换为序数（用于连续性检测）。

    支持格式：
      - 2026-04-15, 2026/04/15
      - 04-15, 4/15
      - 20260415
      - 2026年4月15日
    """
    if not date_str or not isinstance(date_str, str):
        return None
    date_str = date_str.strip()

    # 尝试多种格式
    formats = [
        (r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', lambda m: (int(m.group(1)), int(m.group(2)), int(m.group(3)))),
        (r'^(\d{4})年(\d{1,2})月(\d{1,2})日$', lambda m: (int(m.group(1)), int(m.group(2)), int(m.group(3)))),
        (r'^(\d{1,2})[-/](\d{1,2})$', lambda m: (datetime.now().year, int(m.group(1)), int(m.group(2)))),
        (r'^(\d{8})$', lambda m: (int(m.group(1)[:4]), int(m.group(1)[4:6]), int(m.group(1)[6:8]))),
    ]

    for pattern, extractor in formats:
        m = re.match(pattern, date_str)
        if m:
            try:
                y, mo, d = extractor(m)
                return datetime(y, mo, d).toordinal()
            except (ValueError, OverflowError):
                return None

    return None


def _ordinal_to_date_str(ordinal: int) -> str:
    """将序数转换为日期字符串 YYYY-MM-DD。"""
    dt = datetime.fromordinal(ordinal)
    return dt.strftime("%Y-%m-%d")


# ========== 文档关联图谱 ==========
def build_link_graph(index: Dict[str, Any], out_base: Path) -> Dict[str, Any]:
    """Build document link graph for cross-document audit task loading.

    Creates edges between documents that share:
    - same_pile: same pile number across docs
    - same_date_log: same construction date
    - signer_in_roster: signer appears in personnel roster
    """
    nodes = {}
    edges = []
    docs = index.get("documents", [])

    for doc in docs:
        doc_id = doc.get("id", "")
        if not doc_id:
            continue
        data_file = doc.get("data_file")
        nodes[doc_id] = {
            "file": data_file,
            "doc_type": doc.get("doc_type", ""),
            "professional": doc.get("professional", ""),
            "subdivision": doc.get("subdivision_code"),
            "key_fields": {},
        }
        # Load structured data to extract key fields
        if data_file:
            data_path = out_base / data_file
            if data_path.exists():
                try:
                    data = json.loads(data_path.read_text(encoding="utf-8"))
                    rows = data.get("structured_rows") or data.get("rows", [])
                    if rows:
                        first_row = rows[0]
                        nodes[doc_id]["key_fields"] = {
                            k: str(v) for k, v in list(first_row.items())[:10]
                        }
                except Exception:
                    pass

    # Build edges by comparing key fields
    doc_ids = list(nodes.keys())
    for i, id_a in enumerate(doc_ids):
        node_a = nodes[id_a]
        keys_a = node_a.get("key_fields", {})
        for id_b in doc_ids[i+1:]:
            node_b = nodes[id_b]
            keys_b = node_b.get("key_fields", {})

            # Check same pile number
            pile_a = keys_a.get("pile_no", "")
            pile_b = keys_b.get("pile_no", "")
            if pile_a and pile_b and pile_a == pile_b:
                edges.append({
                    "from": id_a, "to": id_b,
                    "type": "same_pile",
                    "strength": 1.0,
                    "join_key": {"pile_no": pile_a},
                    "rule_hint": "施工记录实长↔检验批验收长度"
                })

            # Check same date
            date_a = keys_a.get("施工日期", "") or keys_a.get("date", "")
            date_b = keys_b.get("施工日期", "") or keys_b.get("date", "")
            if date_a and date_b and date_a == date_b:
                edges.append({
                    "from": id_a, "to": id_b,
                    "type": "same_date_log",
                    "strength": 0.8,
                    "join_key": {"date": date_a},
                    "rule_hint": "施工记录↔施工日志每日合计对照"
                })

    return {
        "schema_version": "1.0",
        "built_at": now_iso(),
        "node_count": len(nodes),
        "edge_count": len(edges),
        "nodes": nodes,
        "edges": edges,
    }


# ========== 主流程 ==========
def main() -> int:
    parser = argparse.ArgumentParser(
        description="建立民航施工资料数据底座（Phase 1）"
    )
    parser.add_argument("project_path", help="项目文件夹路径")
    parser.add_argument(
        "--engine", choices=["auto", "vision", "paddle"], default="auto",
        help="OCR 引擎（默认 auto）"
    )
    parser.add_argument(
        "--incremental", action="store_true",
        help="增量模式：仅处理新文件或已变更文件"
    )
    parser.add_argument(
        "--out", default="数据底座",
        help="数据底座目录名（默认：数据底座）"
    )
    parser.add_argument(
        "--preconditions",
        help="前置信息 JSON 文件路径（含 stage/nature/scope/ocr_engine/special_notes/excluded_files/expected_rows）。"
             "未提供时将在输出目录生成默认前置信息文件。"
    )
    parser.add_argument(
        "--expected-rows",
        help="预期行数 JSON 文件路径，格式：{\"文件名模式\": 行数}"
    )
    parser.add_argument(
        "--include-all", action="store_true",
        help="扫描所有文件扩展名（默认仅扫描资料类扩展名）"
    )
    args = parser.parse_args()

    project_path = Path(args.project_path).resolve()
    if not project_path.is_dir():
        print(f"❌ 项目文件夹不存在: {project_path}", file=sys.stderr)
        return 1

    out_base = project_path / args.out
    out_base.mkdir(parents=True, exist_ok=True)

    # 读取前置信息；未提供时生成默认值并保存到输出目录
    if args.preconditions:
        preconditions_path = Path(args.preconditions).resolve()
        if not preconditions_path.exists():
            print(f"❌ 前置信息文件不存在: {preconditions_path}", file=sys.stderr)
            return 1
        preconditions = load_json(preconditions_path) or {}
    else:
        preconditions = generate_default_preconditions(project_path)

    # 命令行指定的 OCR 引擎优先级最高，回写到前置信息
    preconditions["ocr_engine"] = args.engine

    if not args.preconditions:
        default_pre_path = out_base / "preconditions_default.json"
        save_json(default_pre_path, preconditions)
        print(f"  [i] 未提供前置信息文件，已生成默认值: {default_pre_path}", file=sys.stderr)

    # 读取预期行数配置
    expected_map: Dict[str, int] = {}
    if args.expected_rows:
        ep = load_json(Path(args.expected_rows).resolve()) or {}
        if isinstance(ep, dict):
            expected_map = {k: int(v) for k, v in ep.items()}
    # 也支持从 preconditions.expected_rows 读取
    if isinstance(preconditions.get("expected_rows"), dict):
        for k, v in preconditions["expected_rows"].items():
            expected_map.setdefault(k, int(v))

    excluded_set = set(preconditions.get("excluded_files", []))

    print(f"\n🏗️  开始建立数据底座: {project_path}", file=sys.stderr)
    print(f"   输出目录: {out_base}\n", file=sys.stderr)

    # 初始化/读取 index.json
    index = ensure_index(project_path, out_base, preconditions)
    # 同步前置信息（CLI 引擎优先，覆盖旧值）
    index["preconditions"] = preconditions

    # 扫描文件
    rel_files = scan_files(project_path, args.out, args.include_all)
    print(f"📁 共扫描到 {len(rel_files)} 个文件", file=sys.stderr)

    # 重置分类清单（将按本次扫描结果重新写入）
    file_classification = index.setdefault("file_classification", {
        "audited_files": [],
        "reference_files": [],
        "excluded_files": [],
    })
    file_classification["audited_files"] = []
    file_classification["reference_files"] = []
    file_classification["excluded_files"] = []

    # 预处理：判断是否存在正式施工记录/检验批，用于施工日志自动降级
    has_formal_records = _has_formal_records(rel_files)
    if has_formal_records:
        print(f"  [i] 检测到正式施工记录/检验批，施工日志将自动归类为依据文件", file=sys.stderr)

    # 逐个处理
    for rel in rel_files:
      try:
        abs_path = project_path / rel
        # v7.2 C1: 文件文本预览（前几行）供 LLM 分类语义判定
        text_preview = _read_text_preview(abs_path)
        classification, professional, subcategory, doc_type, classification_source, classification_confidence = classify_file(
            rel, excluded_set, has_formal_records, text_preview=text_preview
        )

        # v7.2 C2: 图纸角色解耦——is_drawing 标签与 doc_role 正交
        is_drawing = detect_is_drawing(rel.name)
        drawing_type = detect_drawing_type(rel.name)
        project_stage = preconditions.get("stage", "分部分项验收")
        doc_role = infer_doc_role(is_drawing, project_stage, classification, drawing_type)

        if classification == "excluded_files":
            file_classification["excluded_files"].append(str(rel))
            print(f"  [排除] {rel.name}", file=sys.stderr)
            continue

        # 图纸 reference 角色：作为依据文件，只记录元数据 doc 条目，不 OCR 提取
        if is_drawing and doc_role == "reference":
            file_classification["reference_files"].append(str(rel))
            print(f"  [图纸·依据] {rel.name}", file=sys.stderr)
            content_hash = _compute_file_hash(abs_path)
            ref_doc = {
                "id": next_doc_id(index),
                "original_file": str(rel),
                "file_type": Path(rel.name).suffix.upper().lstrip(".") or "UNKNOWN",
                "is_scanned": False,
                "extraction_mode": "reference_skip",
                "is_drawing": True,
                "drawing_type": drawing_type,
                "doc_role": "reference",
                "doc_type": doc_type,
                "professional": professional,
                "subcategory": subcategory,
                "classification_source": classification_source,
                "classification_confidence": classification_confidence,
                "subdivision_code": None,
                "pages": 1,
                "ocr_status": "skipped",
                "ocr_engine": "",
                "ocr_confidence": 0.0,
                "data_file": None,
                "human_verified": True,
                "human_confirmed": True,   # v7.2 C1: 依据文件角色由规则确定，无需 AI 分类确认
                "audit_status": "skipped",
                "last_updated": now_iso(),
                "size_bytes": abs_path.stat().st_size,
                "content_hash": content_hash,
            }
            update_index_for_doc(index, ref_doc)
            continue

        # 图纸 audited 角色：需审核，继续走提取流程
        if is_drawing:
            file_classification["audited_files"].append(str(rel))
            print(f"  [图纸·审核] {rel.name}", file=sys.stderr)
        elif classification == "reference_files":
            file_classification["reference_files"].append(str(rel))
            print(f"  [依据] {rel.name}", file=sys.stderr)
            continue
        else:
            file_classification["audited_files"].append(str(rel))

        # 格式识别
        sniff = sniff_document(str(abs_path))
        pages = sniff.get("page_count") or 1
        is_scanned = bool(sniff.get("is_scanned"))
        method = sniff.get("extraction_method", "unknown")
        file_type = sniff.get("suffix", "").upper().lstrip(".") or "UNKNOWN"
        if file_type in ("PNG", "JPG", "JPEG", "BMP", "TIFF", "TIF"):
            file_type = "IMAGE"

        # v7.2 C3: extraction_mode 字段——区分数据提取方式，避免电子表谎报 OCR%
        _sniff_method = sniff.get("extraction_method", "unknown")
        if file_type == "IMAGE":
            extraction_mode = "image"
        elif _sniff_method == "ocr":
            extraction_mode = "ocr"
        elif _sniff_method == "pymupdf":
            extraction_mode = "text_pdf"
        elif _sniff_method == "docx":
            extraction_mode = "docx"
        elif _sniff_method == "excel":
            extraction_mode = "meta_xlsx"
        elif _sniff_method == "text":
            extraction_mode = "text_pdf"
        else:
            extraction_mode = "unknown"

        print(f"\n📄 {rel.name} | type={file_type} | pages={pages} | scanned={is_scanned} | method={method} | mode={extraction_mode}", file=sys.stderr)

        # 增量模式：已存在且未变更则跳过
        existing_doc = find_doc_by_original(index, str(rel))
        content_hash = _compute_file_hash(abs_path)
        if args.incremental and existing_doc:
            old_hash = existing_doc.get("content_hash", "")
            if existing_doc.get("ocr_status") == "completed" and old_hash and old_hash == content_hash:
                print(f"  [增量跳过] 已处理且内容未变更", file=sys.stderr)
                continue
            elif old_hash and old_hash != content_hash:
                print(f"  [增量更新] 文件内容已变更，重新处理", file=sys.stderr)

        # 确定存储目录与文件名
        folder = out_base / safe_filename(professional) / safe_filename(subcategory)
        folder.mkdir(parents=True, exist_ok=True)
        base_name = safe_filename(abs_path.stem)

        data_file = folder / f"{base_name}.json"
        md_file = folder / f"{base_name}.md"
        ocr_raw_file = folder / f"{base_name}_ocr.json"
        quality_file = folder / f"{base_name}_quality.json"
        confusion_file = folder / f"{base_name}_confusion.json"
        text_file = folder / f"{base_name}_raw.txt"

        ocr_status = "completed"
        ocr_engine = "unknown"
        ocr_confidence = 0.0
        ocr_text = ""
        retry_log: List[Dict[str, Any]] = []

        # ===== 提取 =====
        if method == "ocr":
            # 扫描件 PDF / 图片
            engine = args.engine
            result = call_ocr_image(abs_path, engine, text_file, ocr_raw_file)
            ocr_text = result.get("text", "")
            ocr_engine = result.get("engine", engine)
            ocr_confidence = result.get("confidence", 0.0)

            # 行数校验与自动重试（铁律 R-16）
            rows = build_rows(ocr_text, doc_type)
            expected = get_expected_rows(rel, pages, file_type, expected_map)
            passed, ratio = check_row_count(len(rows), expected)

            if not passed:
                # 第一次重试：图像增强（增量模式也执行，确保数据质量）
                print(f"  [!] 行数不足（实际 {len(rows)} / 预期 {expected}，比例 {ratio:.0%}），尝试增强重识别...", file=sys.stderr)
                retry_log.append({"attempt": 1, "action": "preprocess_enhance", "engine": engine})
                result = call_ocr_image(abs_path, engine, text_file, ocr_raw_file, preprocess="enhance")
                ocr_text = result.get("text", "")
                ocr_engine = result.get("engine", engine)
                ocr_confidence = result.get("confidence", 0.0)
                rows = build_rows(ocr_text, doc_type)
                passed, ratio = check_row_count(len(rows), expected)

                # 第二次重试：切换引擎（paddle → vision）
                if not passed:
                    fallback_engine = "vision" if engine in ("paddle", "auto") else "vision"
                    print(f"  [!] 增强后仍不足，尝试切换引擎 {fallback_engine}...", file=sys.stderr)
                    retry_log.append({"attempt": 2, "action": "switch_engine", "from": engine, "to": fallback_engine})
                    result = call_ocr_image(abs_path, fallback_engine, text_file, ocr_raw_file)
                    if result.get("text"):
                        ocr_text = result.get("text", "")
                        ocr_engine = result.get("engine", fallback_engine)
                        ocr_confidence = result.get("confidence", 0.0)
                        rows = build_rows(ocr_text, doc_type)
                        passed, ratio = check_row_count(len(rows), expected)

                if not passed:
                    ocr_status = "needs_review"
                    retry_log.append({
                        "attempt": "final",
                        "action": "mark_needs_review",
                        "reason": f"实际行数 {len(rows)} 低于预期 {expected} 的 {ROW_COUNT_THRESHOLD:.0%}",
                    })
                    print(f"  [⚠️] 行数校验未通过，标记为 needs_review", file=sys.stderr)

        elif method == "pymupdf":
            # 电子档 PDF
            result = call_extract_pdf(abs_path, text_file)
            ocr_text = result.get("text", "")
            ocr_engine = result.get("engine", "PyMuPDF")
            ocr_confidence = result.get("confidence", 1.0)
            # 写入 ocr_raw_file（仅做追溯用）
            save_json(ocr_raw_file, {
                "text": ocr_text,
                "engine": ocr_engine,
                "confidence": ocr_confidence,
                "items": [],
                "page_count": pages,
                "source": "extract_pdf.py",
            })

        elif method == "text":
            # 纯文本文件
            result = call_extract_text(abs_path, text_file)
            ocr_text = result.get("text", "")
            ocr_engine = result.get("engine", "text")
            ocr_confidence = result.get("confidence", 1.0)
            save_json(ocr_raw_file, {
                "text": ocr_text,
                "engine": ocr_engine,
                "confidence": ocr_confidence,
                "items": [],
                "page_count": pages,
                "source": "direct_read",
            })

        elif method == "excel":
            # Excel 文件：提取为结构化 rows
            try:
                import openpyxl
                wb = openpyxl.load_workbook(abs_path, data_only=True)
                sheet_texts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    sheet_rows = []
                    for row in ws.iter_rows(values_only=True):
                        # 跳过全空行
                        if any(v is not None and str(v).strip() for v in row):
                            sheet_rows.append([str(v) if v is not None else "" for v in row])
                    if sheet_rows:
                        sheet_texts.append(f"--- Sheet: {sheet_name} ---")
                        for row in sheet_rows:
                            sheet_texts.append("\t".join(row))
                wb.close()
                ocr_text = "\n".join(sheet_texts)
                ocr_engine = "openpyxl"
                ocr_confidence = 1.0
                pages = len(wb.sheetnames)
                save_json(ocr_raw_file, {
                    "text": ocr_text,
                    "engine": ocr_engine,
                    "confidence": ocr_confidence,
                    "items": [],
                    "page_count": pages,
                    "source": "openpyxl",
                    "sheets": wb.sheetnames,
                })
            except ImportError as e:
                ocr_status = "unsupported"
                ocr_engine = "excel"
                ocr_confidence = 0.0
                reason = "openpyxl 未安装，无法读取 Excel"
                save_json(ocr_raw_file, {
                    "text": "",
                    "engine": "excel",
                    "confidence": 0.0,
                    "items": [],
                    "reason": reason,
                })
                print(f"  [!] {reason}", file=sys.stderr)
            except Exception as e:
                ocr_status = "needs_review"
                ocr_engine = "excel"
                ocr_confidence = 0.0
                reason = f"Excel 读取异常: {e}"
                save_json(ocr_raw_file, {
                    "text": "",
                    "engine": "excel",
                    "confidence": 0.0,
                    "items": [],
                    "reason": reason,
                })
                print(f"  [!] {reason}", file=sys.stderr)

        else:
            # 暂不支持的类型
            ocr_status = "unsupported"
            ocr_engine = method
            ocr_confidence = 0.0
            reason = f"暂不支持的提取方式: {method}"
            save_json(ocr_raw_file, {
                "text": "",
                "engine": method,
                "confidence": 0.0,
                "items": [],
                "reason": reason,
            })
            print(f"  [!] {reason}", file=sys.stderr)

        if method == "excel":
            # Excel: only record metadata in index, don't create data files
            doc = {
                "id": next_doc_id(index),
                "original_file": str(rel),
                "file_type": "XLSX",
                "is_scanned": False,
                "extraction_mode": extraction_mode,
                "is_drawing": is_drawing,
                "drawing_type": drawing_type,
                "doc_role": doc_role,
                "doc_type": doc_type,
                "professional": professional,
                "subcategory": subcategory,
                "classification_source": classification_source,
                "classification_confidence": classification_confidence,
                "subdivision_code": None,
                "pages": len(wb.sheetnames) if 'wb' in dir() else 1,
                "ocr_status": "completed",
                "ocr_engine": "openpyxl",
                "ocr_confidence": 1.0,
                "data_file": None,
                "data_md": None,
                "data_format": "excel_raw",
                "metadata": {
                    "sheets": wb.sheetnames if 'wb' in dir() else [],
                    "total_rows": len(sheet_rows) if 'sheet_rows' in dir() else 0,
                },
                "human_verified": True,
                "human_confirmed": False,  # v7.2 C1: AI 分类结果，待人工确认
                "audit_status": "pending",
                "last_updated": now_iso(),
                "size_bytes": sniff.get("size_bytes"),
                "content_hash": content_hash,
            }
            update_index_for_doc(index, doc)
            print(f"  ✓ Excel元数据已记录: {rel.name}", file=sys.stderr)
            continue  # Skip the rest of the loop for Excel files

        # ===== 生成结构化 rows =====
        rows: List[Dict[str, Any]] = []
        if ocr_status != "unsupported":
            rows = build_rows(ocr_text, doc_type)
            # 空内容也触发 needs_review
            if not rows:
                ocr_status = "needs_review"
                print(f"  [!] 未识别到任何数据行", file=sys.stderr)

        # 构建 page_map（三层结构之 Layer 3）
        page_map = build_page_map(abs_path, is_scanned, ocr_text, method)

        # v7.2 C5: 表头三路融合——记录原始表头和映射，支持人工确认与自成长
        header_info = extract_header_info(ocr_text, doc_type)

        structured = {
            "schema_version": "1.0",
            "doc_id": existing_doc.get("id") if existing_doc else next_doc_id(index),
            "doc_type": doc_type,
            "source_file": str(rel),
            "professional": professional,
            "subcategory": subcategory,
            "ocr_engine": ocr_engine,
            "ocr_confidence": ocr_confidence,
            "ocr_completed_at": now_iso(),
            "structured_rows": rows,                                  # Layer 1: for rule engine
            "full_text": ocr_text,                                    # Layer 2: complete text for LLM
            "page_map": page_map,                                     # Layer 3: per-page text + image refs
            "fields_detected": list(rows[0].keys()) if rows else [],
            "rows": rows,                                             # backward-compat alias
            "quality_result": {},
            "confusion_result": {},
            "corrections_applied": [],
            # v7.2 C5: 表头识别元数据
            "raw_headers": header_info["raw_headers"],
            "header_mapping": header_info["header_mapping"],
            "header_confidence": header_info["header_confidence"],
            "header_needs_human_confirm": header_info["needs_human_confirm"],
            "header_math_chain_verified": header_info["math_chain_verified"],
            "header_math_chain_failures": header_info["math_chain_failures"],
            "header_math_chain_inferred": header_info["math_chain_inferred"],
            "header_column_feature_issues": header_info["column_feature_issues"],
        }

        # ===== 分配分部分项 code（v6.0 新增） =====
        subdivision_code = assign_subdivision_to_document(
            professional=professional,
            doc_type=doc_type,
            subcategory=subcategory,
        )
        if subdivision_code:
            sub_info = get_subdivision_info(professional, subdivision_code)
            structured["subdivision_code"] = subdivision_code
            structured["subdivision_label"] = sub_info["label"] if sub_info else ""
            structured["sub_division"] = sub_info["sub_label"] if sub_info else ""
            print(f"  [分部分项] {subdivision_code} {structured['subdivision_label']}（{structured['sub_division']}）", file=sys.stderr)
        else:
            structured["subdivision_code"] = None
            structured["subdivision_label"] = ""
            structured["sub_division"] = ""
            print(f"  [!] 未能匹配分部分项", file=sys.stderr)
        save_json(data_file, structured)

        # ===== 数据质量检测 & 混淆检测 =====
        quality_alerts = 0
        confusion_suspects = 0
        if ocr_status not in ("unsupported",):
            quality_result, quality_alerts = call_data_quality_check(data_file)
            save_json(quality_file, quality_result)
            confusion_result, confusion_suspects = call_ocr_confusion_check(data_file)
            save_json(confusion_file, confusion_result)

            # 将质量/混淆结果回填到结构化 JSON
            structured["quality_result"] = quality_result
            structured["confusion_result"] = confusion_result
            save_json(data_file, structured)

        # ===== 更新 index.json =====
        doc = {
            "id": structured["doc_id"],
            "original_file": str(rel),
            "file_type": file_type,
            "is_scanned": is_scanned,
            "extraction_mode": extraction_mode,
            "is_drawing": is_drawing,
            "drawing_type": drawing_type,
            "doc_role": doc_role,
            "doc_type": doc_type,
            "professional": professional,
            "subcategory": subcategory,
            "classification_source": classification_source,
            "classification_confidence": classification_confidence,
            "subdivision_code": structured.get("subdivision_code"),
            "subdivision_label": structured.get("subdivision_label", ""),
            "sub_division": structured.get("sub_division", ""),
            "pages": pages,
            "ocr_status": ocr_status,
            "ocr_engine": ocr_engine,
            "ocr_confidence": ocr_confidence,
            "ocr_completed_at": structured["ocr_completed_at"],
            "data_file": str(data_file.relative_to(out_base)).replace("\\", "/"),
            "data_md": None,
            "ocr_raw_file": str(ocr_raw_file.relative_to(out_base)).replace("\\", "/"),
            "quality_file": str(quality_file.relative_to(out_base)).replace("\\", "/"),
            "confusion_file": str(confusion_file.relative_to(out_base)).replace("\\", "/"),
            "quality_alerts": quality_alerts,
            "confusion_suspects": confusion_suspects,
            "human_verified": not (is_scanned or file_type == "IMAGE"),
            "human_confirmed": False,  # v7.2 C1: AI 分类结果，待人工确认
            "corrected_file": None,
            "audit_status": "pending",
            "last_updated": now_iso(),
            "size_bytes": sniff.get("size_bytes"),
            "content_hash": content_hash,
            "retry_log": retry_log,
        }
        update_index_for_doc(index, doc)
        print(f"  ✓ 已生成: {data_file.name}, 告警 {quality_alerts}, 存疑 {confusion_suspects}", file=sys.stderr)
      except Exception as e:
        print(f"  ❌ 处理失败，跳过此文件: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        continue

    # 同步 documents 与当前 file_classification，清理 stale 条目
    sync_documents_with_classification(index)

    # 断档检测（N-09）：桩号/日期/编号连续性
    print(f"\n🔍 开始断档检测...", file=sys.stderr)
    detect_gaps(index, out_base)

    # ========== Build link graph ==========
    print(f"\n🔗 构建文档关联图谱...", file=sys.stderr)
    link_graph = build_link_graph(index, out_base)
    save_json(out_base / "link_graph.json", link_graph)

    # 更新 index 元信息
    index["updated_at"] = now_iso()
    index["stage"] = "foundation_built"
    # v7.2 C1/C2: C-01 人工分类确认闸门——分类未经人工确认前为 false，
    # data-editor 文档属性面板确认分类后置 true，审核阶段检查
    index["file_classification_confirmed"] = False
    index["classification_pending_count"] = sum(
        1 for d in index.get("documents", [])
        if d.get("doc_role", "audited") != "reference" and d.get("human_confirmed") is not True
    )
    save_json(out_base / "index.json", index)

    # 复制 Web 模板
    copy_web_templates(project_path)

    print(f"\n✅ 数据底座建立完成: {out_base}", file=sys.stderr)
    print(f"   阶段: {index['stage']}", file=sys.stderr)
    print(f"   被审核文件: {len(file_classification['audited_files'])}", file=sys.stderr)
    print(f"   依据文件: {len(file_classification['reference_files'])}", file=sys.stderr)
    print(f"   排除文件: {len(file_classification['excluded_files'])}", file=sys.stderr)

    # v7.2 C2: 竣工图纸高亮提示
    drawings = [d for d in index.get("documents", []) if d.get("is_drawing")]
    if drawings:
        completion_drawings = [d for d in drawings if d.get("doc_role") == "audited"]
        if completion_drawings:
            print(f"\n🟡 竣工图纸提醒：检测到 {len(completion_drawings)} 份图纸在竣工阶段项目中需审核。", file=sys.stderr)
            print(f"   这些图纸默认角色为「审核文件」，请通过 C-01 人工确认闸门确认角色：", file=sys.stderr)
            for d in completion_drawings:
                print(f"   - {d.get('original_file', d.get('id', '?'))}", file=sys.stderr)
        else:
            print(f"\n[i] 检测到 {len(drawings)} 份图纸（施工阶段，默认作为依据文件）。", file=sys.stderr)
    # Check if any documents need manual verification
    unverified = [d for d in index.get("documents", []) if not d.get("human_verified", False)]
    if unverified:
        print(f"\n⛔ Phase 1 结束。{len(unverified)} 份扫描件需要人工核对。", file=sys.stderr)
        print(f"   请打开项目文件夹中的 data-editor.html 完成人工核对后再继续审核。", file=sys.stderr)
        for d in unverified:
            print(f"   - {d.get('original_file', d.get('id', '?'))}", file=sys.stderr)
    else:
        print(f"\n✅ Phase 1 结束。所有文档已自动通过，可直接进入审核阶段。", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
